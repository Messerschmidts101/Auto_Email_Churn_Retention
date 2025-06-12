from pyspark.sql.types import StructType, StructField, StringType, DateType, IntegerType, ArrayType, BooleanType
from pyspark.sql import DataFrame
import pyspark.sql.functions as F
import pyspark.sql.window as W
from databricks import automl
import mlflow
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from itertools import chain 
from datetime import datetime
import pandas as pd
import time
import io


class utils:
    def __init__(self):
            self.dicParamsDefault = {
            'Primary_Keys': ['tblService_SEHIID','tblService_BILLD','tblService_WRKORDNO','tblService_VEHIID','tblService_CUSTNO','tblCustomer_CUSTNO','tblService_DPCUSTNO','tblService_HCUSTNO','tblService_OCUSTNO'],
            'Car_Makes':['jaguar','land rover','rover','range rover','bmw'],
            'Evaluation_Metric':'f1',
            'Positive_Label':1,
            'DATAROBOT_API_KEY':None, # deprecated
            'DATAROBOT_ENDPOINT':None, # deprecated
            'DR_Project_Name':'Churn Prediction Latvia BMW & JLR', # deprecated
            'intServiceWindowCarAge': 365,
            'intServiceWindowMileage': 6000,
            'lisintExcludedMileageValues': [0,1],
            'lisstrExcludedCustomerGroup': ['AN', 'V', 'VBMA', 'VBML', 'I'],
            'lisstrIncludedCustomerGroup':['A'],
            'intLocalMinimumFirst_Service_Grouping': 4,
            'intLocalMinimumSecond_Service_Grouping': 12,
            'intDefault_Service_Grouping':15,
            'fltProportionThreshold': 0.1,
            'lisstrModellingFeatures':[
                # Normal
                'new_Churn_Flag',
                'new_Make_Model',
                'tblVehicle_MAKE',
                'tblVehicle_MCODE',
                'tblVehicle_MYEAR',
                'tblVehicle_PURCHPR',
                'tblService_DISTDRIV',
                'new_TSUM_Total',
                'new_TSUM_Total_Previous',
                'new_TSUM_Coverage_Personal',
                'new_TSUM_Coverage_Warranty',
                'new_Service_Group',
                'new_Age_Of_Car',
                'new_Age_Of_Car_Previous',
                'new_Expected_Vs_Actual_Visits_By_Age_Of_Car',
                'new_Expected_Vs_Actual_Visits_By_DISTDRIV',
                'new_Days_Passed_Previous_Service',
                'new_Mileage_Passed_Previous_Service',
                'new_Warranty_Period_Left_In_Days',
                'new_Warranty_Period_Left_By_Mileage',
                'new_Economy_2008_Financial_Crisis',
                'new_Economy_2020_Covid_19',
                'new_Economy_2022_Ukraine_Russia_War',
                'new_Service_Duration',
                'new_Service_Duration_Previous',
                'new_GROWID_Count',
                'new_Churn_Flag_Previous',
                # Rolling_Average 0
                'new_Days_Passed_Previous_Service_Rolling_0_Service_Average',
                'new_Mileage_Passed_Previous_Service_Rolling_0_Service_Average',
                'new_GROWID_Count_Rolling_0_Service_Average',
                'new_Service_Duration_Rolling_0_Service_Average',
                'new_TSUM_Coverage_Personal_Rolling_0_Service_Average',
                'new_TSUM_Coverage_Warranty_Rolling_0_Service_Average',
                'new_TSUM_Coverage_Insurance_Rolling_0_Service_Average',
                'new_TSUM_Coverage_Others_Rolling_0_Service_Average',
                'new_TSUM_Total_Rolling_0_Service_Average',
                # Rolling_Average 3
                'new_Days_Passed_Previous_Service_Rolling_3_Service_Average',
                'new_Mileage_Passed_Previous_Service_Rolling_3_Service_Average',
                'new_GROWID_Count_Rolling_3_Service_Average',
                'new_Service_Duration_Rolling_3_Service_Average',
                'new_TSUM_Coverage_Personal_Rolling_3_Service_Average',
                'new_TSUM_Coverage_Warranty_Rolling_3_Service_Average',
                'new_TSUM_Coverage_Insurance_Rolling_3_Service_Average',
                'new_TSUM_Coverage_Others_Rolling_3_Service_Average',
                'new_TSUM_Total_Rolling_3_Service_Average',
                # Rolling_Sum 0 
                'new_Churn_Flag_Rolling_0_Service_Sum',
                'new_TSUM_Coverage_Personal_Rolling_0_Service_Sum',
                'new_TSUM_Coverage_Warranty_Rolling_0_Service_Sum',
                'new_TSUM_Coverage_Insurance_Rolling_0_Service_Sum',
                'new_TSUM_Coverage_Others_Rolling_0_Service_Sum',
                'new_TSUM_Total_Rolling_0_Service_Sum',
                # Rolling_Sum 3
                'new_Churn_Flag_Rolling_3_Service_Sum',
                'new_TSUM_Coverage_Personal_Rolling_3_Service_Sum',
                'new_TSUM_Coverage_Warranty_Rolling_3_Service_Sum',
                'new_TSUM_Coverage_Insurance_Rolling_3_Service_Sum',
                'new_TSUM_Coverage_Others_Rolling_3_Service_Sum',
                'new_TSUM_Total_Rolling_3_Service_Sum',
                # Rolling_Max 0
                'new_Churn_Flag_Rolling_0_Service_Max',
                'new_Days_Passed_Previous_Service_Rolling_0_Service_Max',
                'new_Mileage_Passed_Previous_Service_Rolling_0_Service_Max',
                'new_Service_Duration_Rolling_0_Service_Max',
                'new_TSUM_Coverage_Personal_Rolling_0_Service_Max',
                'new_TSUM_Coverage_Warranty_Rolling_0_Service_Max',
                'new_TSUM_Coverage_Insurance_Rolling_0_Service_Max',
                'new_TSUM_Coverage_Others_Rolling_0_Service_Max',
                'new_TSUM_Total_Rolling_0_Service_Max',
                # Rolling_Max 3
                'new_Churn_Flag_Rolling_3_Service_Max',
                'new_Days_Passed_Previous_Service_Rolling_3_Service_Max',
                'new_Mileage_Passed_Previous_Service_Rolling_3_Service_Max',
                'new_Service_Duration_Rolling_3_Service_Max',
                'new_TSUM_Coverage_Personal_Rolling_3_Service_Max',
                'new_TSUM_Coverage_Warranty_Rolling_3_Service_Max',
                'new_TSUM_Coverage_Insurance_Rolling_3_Service_Max',
                'new_TSUM_Coverage_Others_Rolling_3_Service_Max',
                'new_TSUM_Total_Rolling_3_Service_Max',
                # Difference From Rolling_Average 0
                'new_Days_Passed_Previous_Service_Difference_From_Rolling_0_Service_Average',
                'new_Mileage_Passed_Previous_Service_Difference_From_Rolling_0_Service_Average',
                'new_GROWID_Count_Difference_From_Rolling_0_Service_Average',
                'new_Service_Duration_Difference_From_Rolling_0_Service_Average',
                'new_TSUM_Coverage_Personal_Difference_From_Rolling_0_Service_Average',
                'new_TSUM_Coverage_Warranty_Difference_From_Rolling_0_Service_Average',
                'new_TSUM_Coverage_Insurance_Difference_From_Rolling_0_Service_Average',
                'new_TSUM_Coverage_Others_Difference_From_Rolling_0_Service_Average',
                'new_TSUM_Total_Difference_From_Rolling_0_Service_Average',
                # Difference From Rolling_Average 3
                'new_Days_Passed_Previous_Service_Difference_From_Rolling_3_Service_Average',
                'new_Mileage_Passed_Previous_Service_Difference_From_Rolling_3_Service_Average',
                'new_GROWID_Count_Difference_From_Rolling_3_Service_Average',
                'new_Service_Duration_Difference_From_Rolling_3_Service_Average',
                'new_TSUM_Coverage_Personal_Difference_From_Rolling_3_Service_Average',
                'new_TSUM_Coverage_Warranty_Difference_From_Rolling_3_Service_Average',
                'new_TSUM_Coverage_Insurance_Difference_From_Rolling_3_Service_Average',
                'new_TSUM_Coverage_Others_Difference_From_Rolling_3_Service_Average',
                'new_TSUM_Total_Difference_From_Rolling_3_Service_Average',
                # Difference From Rolling_Max 0
                'new_Days_Passed_Previous_Service_Difference_From_Rolling_0_Service_Max',
                'new_Mileage_Passed_Previous_Service_Difference_From_Rolling_0_Service_Max',
                'new_Service_Duration_Difference_From_Rolling_0_Service_Max',
                'new_TSUM_Coverage_Personal_Difference_From_Rolling_0_Service_Max',
                'new_TSUM_Coverage_Warranty_Difference_From_Rolling_0_Service_Max',
                'new_TSUM_Coverage_Insurance_Difference_From_Rolling_0_Service_Max',
                'new_TSUM_Coverage_Others_Difference_From_Rolling_0_Service_Max',
                'new_TSUM_Total_Difference_From_Rolling_0_Service_Max',
                # Difference From Rolling_Max 3
                'new_Days_Passed_Previous_Service_Difference_From_Rolling_3_Service_Max',
                'new_Mileage_Passed_Previous_Service_Difference_From_Rolling_3_Service_Max',
                'new_Service_Duration_Difference_From_Rolling_3_Service_Max',
                'new_TSUM_Coverage_Personal_Difference_From_Rolling_3_Service_Max',
                'new_TSUM_Coverage_Warranty_Difference_From_Rolling_3_Service_Max',
                'new_TSUM_Coverage_Insurance_Difference_From_Rolling_3_Service_Max',
                'new_TSUM_Coverage_Others_Difference_From_Rolling_3_Service_Max',
                'new_TSUM_Total_Difference_From_Rolling_3_Service_Max',
            ],
            'lisstrPresentingFeatures':[
                # obsolete, but remained for future purposes
                'new_Churn_Flag_Previous',
                'tblVehicle_MYEAR',
                'tblVehicle_PURCHPR',
                'tblService_DISTDRIV',
                'new_Make_Model',
                'new_TSUM_Total',
                'new_TSUM_Total_Previous',
                'new_Service_Group',
                'new_Age_Of_Car',
                'new_Age_Of_Car_Previous',
                'new_Expected_Vs_Actual_Visits_By_Age_Of_Car',
                'new_Expected_Vs_Actual_Visits_By_DISTDRIV',
                'new_Days_Passed_Previous_Service',
                'new_Mileage_Passed_Previous_Service',
                'new_Warranty_Period_Left_In_Days',
                'new_Warranty_Period_Left_By_Mileage',
                'new_GROWID_Count',
                'new_Days_Passed_Previous_Service_Rolling_0_Service_Average',
                'new_Service_Duration_Rolling_0_Service_Average',
                'new_TSUM_Coverage_Warranty_Rolling_0_Service_Average',
                'new_TSUM_Total_Rolling_0_Service_Average',
                'new_Churn_Flag_Rolling_0_Service_Sum',
                'new_TSUM_Total_Rolling_0_Service_Sum',
                'new_Churn_Flag_Rolling_3_Service_Sum',
                'new_TSUM_Total_Rolling_0_Service_Max',
                'new_Days_Passed_Previous_Service_Rolling_3_Service_Max',
                'new_Days_Passed_Previous_Service_Difference_From_Rolling_0_Service_Average',
                'new_Mileage_Passed_Previous_Service_Difference_From_Rolling_0_Service_Average',
                'new_Service_Duration_Difference_From_Rolling_0_Service_Average',
                'new_TSUM_Coverage_Personal_Difference_From_Rolling_0_Service_Average',
                'new_TSUM_Coverage_Warranty_Difference_From_Rolling_0_Service_Average',
                'new_Days_Passed_Previous_Service_Difference_From_Rolling_3_Service_Average',
                'new_Mileage_Passed_Previous_Service_Difference_From_Rolling_3_Service_Average',
                'new_Mileage_Passed_Previous_Service_Difference_From_Rolling_0_Service_Max',
            ],
            'lisstrColumnY':['new_Churn_Flag'],
            'tblService_SEHIID':'tblService_SEHIID',
            'tblService_CUSTNO':'tblService_CUSTNO',
            'tblCustomer_CUSTNO':'tblCustomer_CUSTNO',
            'tblCustomer_CGRPID':'tblCustomer_CGRPID',
            'tblService_WRKORDNO':'tblService_WRKORDNO',
            'tblService_BILLD':'tblService_BILLD',
            'tblService_SERVD':'tblService_SERVD',
            'tblService_VEHIID':'tblService_VEHIID',
            'tblVehicle_VEHIID':'tblVehicle_VEHIID',
            'tblVehicle_MAKE':'tblVehicle_MAKE',
            'tblVehicle_MODEL':'tblVehicle_MODEL',
            'tblVehicle_MCODE':'tblVehicle_MCODE',
            'tblVehicle_MYEAR':'tblVehicle_MYEAR',
            'tblVehicle_PURCHPR':'tblVehicle_PURCHPR',
            'tblService_TSUM':'tblService_TSUM',
            'tblService_UNITID':'tblService_UNITID',
            'tblVehicle_FREGD':'tblVehicle_FREGD',
            'tblService_DISTDRIV':'tblService_DISTDRIV',
            'new_Service_Duration_Previous':'new_Service_Duration_Previous',
            'new_Days_Passed_Previous_Service':'new_Days_Passed_Previous_Service',
            'new_Mileage_Passed_Previous_Service':'new_Mileage_Passed_Previous_Service',
            'new_Days_Passed_First_Service':'new_Days_Passed_First_Service',
            'new_Count_Of_Cars_Owned':'new_Count_Of_Cars_Owned',
            'new_Frequency_Count_Of_Cars_Owned':'new_Frequency_Count_Of_Cars_Owned',
            'new_Warranty_Flag':'new_Warranty_Flag',
            'new_Works_Done':'new_Works_Done',
            'new_Service_Duration_by_BILLD':'new_Service_Duration_by_BILLD',
            'new_Service_Duration_by_SERVD':'new_Service_Duration_by_SERVD',
            'new_initial_Service_Group':'new_initial_Service_Group',
            'new_Collected_WRKORDNO':'new_Collected_WRKORDNO',
            'new_Collected_Unique_WRKORDNO':'new_Collected_Unique_WRKORDNO',
            'new_Collected_Unique_WRKORDNO_Count':'new_Collected_Unique_WRKORDNO_Count',
            'new_Previous_Collected_WRKORDNO':'new_Previous_Collected_WRKORDNO',
            'new_Collected_CUSTNO':'new_Collected_CUSTNO',
            'new_Collected_Unique_CUSTNO_Count':'new_Collected_Unique_CUSTNO_Count',
            'new_Previous_Collected_CUSTNO':'new_Previous_Collected_CUSTNO',
            'new_final_Service_Group':'new_final_Service_Group',
            'new_days_between_service_by_BILLD':'new_days_between_service_by_BILLD',
            'new_days_between_service_by_SERVD':'new_days_between_service_by_SERVD',
            'new_WRKORDNO_Count':'new_WRKORDNO_Count',
            'new_WRKORDNO_Mode':'new_WRKORDNO_Mode', 
            'new_CUSTNO_Count':'new_CUSTNO_Count',
            'new_CUSTNO_Mode':'new_CUSTNO_Mode',
            'new_Previous_latest_bill_date':'new_Previous_latest_bill_date',
            'new_Previous_latest_serv_date':'new_Previous_latest_serv_date',
            'new_Warranty_Count':'new_Warranty_Count',
            'new_Warranty_Mode':'new_Warranty_Mode',
            'new_Collected_Warranty':'new_Collected_Warranty',
            'new_Warranty_TSUM_Sum':'new_Warranty_TSUM_Sum',
            'new_Warranty_TSUM_Percentage':'new_Warranty_TSUM_Percentage',
            'new_Warranty_TSUM_Sum_Increment_Absolute':'new_Warranty_TSUM_Sum_Increment_Absolute',
            'new_Warranty_TSUM_Sum_Increment_Relative':'new_Warranty_TSUM_Sum_Increment_Relative',
            'new_TSUM_Sum': 'new_TSUM_Sum',
            'new_TSUM_Average': 'new_TSUM_Average',
            'new_TSUM_Max':'new_TSUM_Max',
            'new_TSUM_Sum_Increment_Absolute':'new_TSUM_Sum_Increment_Absolute',
            'new_TSUM_Sum_Increment_Relative':'new_TSUM_Sum_Increment_Relative',
            'new_Age_Of_Car':'new_Age_Of_Car',
            'new_Age_Of_Car_Previous':'new_Age_Of_Car_Previous',
            'new_Age_Of_Car_Increment_Absolute':'new_Age_Of_Car_Increment_Absolute',
            'new_Age_Of_Car_Increment_Relative':'new_Age_Of_Car_Increment_Relative',
            'new_DISTDRIV_Increment_Absolute':'new_DISTDRIV_Increment_Absolute',
            'new_DISTDRIV_Increment_Relative':'new_DISTDRIV_Increment_Relative',
            'new_Expected_Visits_By_DISTDRIV':'new_Expected_Visits_By_DISTDRIV',
            'new_Expected_Vs_Actual_Visits_By_DISTDRIV':'new_Expected_Vs_Actual_Visits_By_DISTDRIV',
            'new_Expected_Vs_Actual_Visits_By_DISTDRIV_Absolute': 'new_Expected_Vs_Actual_Visits_By_DISTDRIV_Absolute',
            'new_Expected_Vs_Actual_Visits_By_DISTDRIV_Relative': 'new_Expected_Vs_Actual_Visits_By_DISTDRIV_Relative',
            'new_Expected_Visits_By_Age_Of_Car':'new_Expected_Visits_By_Age_Of_Car',
            'new_Expected_Vs_Actual_Visits_By_Age_Of_Car':'new_Expected_Vs_Actual_Visits_By_Age_Of_Car',
            'new_Expected_Vs_Actual_Visits_By_Age_Of_Car_Absolute':'new_Expected_Vs_Actual_Visits_By_Age_Of_Car_Absolute',
            'new_Expected_Vs_Actual_Visits_By_Age_Of_Car_Relative':'new_Expected_Vs_Actual_Visits_By_Age_Of_Car_Relative',
            'new_Days_Passed_Previous_Service':'new_Days_Passed_Previous_Service',
            'new_Days_Passed_First_Service':'new_Days_Passed_First_Service',
            'new_Expected_Visits': 'new_Expected_Visits',
            'new_Visits_Ratio': 'new_Visits_Ratio',
            'new_Economy_2008_Financial_Crisis':'new_Economy_2008_Financial_Crisis',
            'new_Economy_2014_Eurozone_Membership':'new_Economy_2014_Eurozone_Membership',
            'new_Economy_2020_Covid_19':'new_Economy_2020_Covid_19',
            'new_Economy_2022_Ukraine_Russia_War':'new_Economy_2022_Ukraine_Russia_War',
            'new_Economy_2008_Tata_Motors_Acquisition':'new_Economy_2008_Tata_Motors_Acquisition',
            'new_Economy_2017_Brexit':'new_Economy_2017_Brexit',
            'new_BILLD_Next':'new_BILLD_Next',
            'new_BILLD_Next_Days':'new_BILLD_Next_Days',
            'new_SERVD_Next':'new_SERVD_Next',
            'new_SERVD_Next_Days':'new_SERVD_Next_Days',
            'new_Churn_Flag':'new_Churn_Flag',
            'new_Churn_Flag_Previous':'new_Churn_Flag_Previous',
            'tblGROW_GSALID':'tblGROW_GSALID',
            'tblGROW__UNITID':'tblGROW__UNITID',
            'tblGROW_DISCPC':'tblGROW_DISCPC',
            'tblGROW_ITEM':'tblGROW_ITEM',
            'tblGROW_ITEMNO':'tblGROW_ITEMNO',
            'tblGROW_NAME':'tblGROW_NAME',
            'tblGROW_GROWID':'tblGROW_GROWID',
            'tblGROW_is_discounted':'tblGROW_is_discounted',
            'tblGSAL_GSALID':'tblGSAL_GSALID',
            'tblGSAL__UNITID':'tblGSAL__UNITID',
            'tblGSAL_VEHIID':'tblGSAL_VEHIID',
            'tblGSAL_WRKORDNO':'tblGSAL_WRKORDNO',
            'new_SERVD_Start':'new_SERVD_Start',
            'new_SERVD_End':'new_SERVD_End',
            'new_BILLD_Start':'new_BILLD_Start',
            'new_BILLD_End':'new_BILLD_End',
            'new_VEHIID_Collected':'new_VEHIID_Collected',
            'new_CUSTNO_Collected':'new_CUSTNO_Collected',
            'new_SEHIID_Collected':'new_SEHIID_Collected',
            'new_CGRPID_Collected':'new_CGRPID_Collected',
            'new_TSUM_Coverage_Personal':'new_TSUM_Coverage_Personal',
            'new_TSUM_Coverage_Warranty':'new_TSUM_Coverage_Warranty',
            'new_TSUM_Coverage_Insurance':'new_TSUM_Coverage_Insurance',
            'new_TSUM_Coverage_Others':'new_TSUM_Coverage_Others',
            'new_SERVD_Start_Difference_Previous':'new_SERVD_Start_Difference_Previous',
            'new_SERVD_End_Difference_Previous':'new_SERVD_End_Difference_Previous',
            'new_New_Visit_Flag':'new_New_Visit_Flag',
            'new_Service_Group':'new_Service_Group',
            'new_Unique_Visit_Key':'new_Unique_Visit_Key',
            'new_TSUM_Total':'new_TSUM_Total',
            'new_TSUM_Total_Previous':'new_TSUM_Total_Previous',
            'Date_Warranty_Policy_Change': "2022-07-01T00:00:00Z",
            'new_Warranty_Period_In_Days':'new_Warranty_Period_In_Days',
            'new_Warranty_Period_By_Mileage':'new_Warranty_Period_By_Mileage',
            'new_Warranty_Period_Left_In_Days':'new_Warranty_Period_Left_In_Days',
            'new_Warranty_Period_Left_By_Mileage':'new_Warranty_Period_Left_By_Mileage',
            'new_Warranty_Is_Expired':'new_Warranty_Is_Expired',
            'new_TSUM_Coverage_Warranty':'new_TSUM_Coverage_Warranty',
            'new_TSUM_Coverage_Insurance':'new_TSUM_Coverage_Insurance',
            'new_TSUM_Coverage_Warranty_Percentage':'new_TSUM_Coverage_Warranty_Percentage',
            'new_TSUM_Coverage_Insurance_Percentage':'new_TSUM_Coverage_Insurance_Percentage',
            'new_TSUM_Coverage_Combined_Insurance_Warranty':'new_TSUM_Coverage_Combined_Insurance_Warranty',
            'new_Warranty_Will_Expire_Next_Service_By_Days':'new_Warranty_Will_Expire_Next_Service_By_Days',
            'new_Warranty_Will_Expire_Next_Service_By_Mileage':'new_Warranty_Will_Expire_Next_Service_By_Mileage',
            'new_SERVD_Start':'new_SERVD_Start',
            'new_SERVD_End':'new_SERVD_End',
            'new_Service_Duration':'new_Service_Duration',
            'tblGROW_GROWID':'tblGROW_GROWID',
            'new_GROWID_Count':'new_GROWID_Count',
            'new_Make_Model':'new_Make_Model',
            'tblVIN_VEHIID':'tblVIN_VEHIID',
            'tblVIN_BBILLD':'tblVIN_BBILLD',
            'tblVIN_CLOSED':'tblVIN_CLOSED',
            'tblVIN_VINSID':'tblVIN_VINSID',
            'tblVIN_HCUSTNO':'tblVIN_HCUSTNO',
            'tblVehicle_HCUSTNO':'tblVehicle_HCUSTNO',
            'Next_Exp_Service':'Next_Exp_Service',
            'Next_Exp_Service_Days_Left':'Next_Exp_Service_Days_Left',

            }
            
    ##################################################
    #####                                        #####
    #####          [Universal Functions]         #####
    #####   These are supplementary functions    #####
    #####                                        #####
    ##################################################

    # tested and vetted
    def get_parameter_value(self, dicParams, strKey):
        """
        #Inputs:
        1. [dicParams] dictionary of parameters; these parameters intends to overwrite default value for a specific parameter, this can inlcude potentially new feature names, new service interval, etc.
        2. [strKey]: a string search Key to check in the key of the dictionary.
        #Process/Outputs:
        1. Checks the dicParams if a key exists in the dictionary. If it exists (this means the developer will be intentionally overwritting default value for specific parameter), returns the value of the key. If it does not (this means the developer intentionally wants to utilize default value for specific parameter), returns the default value for that parameter/key.
        """
        if type(dicParams) == type(None):
            return self.dicParamsDefault[strKey]
        else:
            if strKey in dicParams:
                return dicParams[strKey]
            else:
                return self.dicParamsDefault[strKey] 
    # tested and vetted
    def find_mode(self,arrInput):
        """
        #Inputs:
        1. [arrInput]: An array-like object from a PySpark row.

        #Process/Outputs:
        1. Identifies the mode value of the input.
        2. This function is intended to be used as a User Defined Function (UDF) in PySpark.
        """
        if arrInput:
            counts = {}
            for elem in arrInput:
                counts[elem] = counts.get(elem, 0) + 1
            return max(counts, key=counts.get)  # Return the element with the highest count
        return None
    # tested and vetted
    def flatten_get_distinct(self, tblInput, lisstrColumnTargets):
        """ 
        #Inputs:
        1. [tblInput]: target table to flatten
        2. [lisstrColumnTargets]: list of string column names of target to flatten.
        #Process/Outputs:
        1. For each column in `lisstrColumnTargets`, this function:
            a. Flattens nested arrays or structures in the column. (Flattens means puts all multi-level arrays to just 1 level)
            b. Removes duplicate values from the flattened array.
        2. Returns the modified DataFrame with the processed columns.
        """
        for strColumn in lisstrColumnTargets:
            tblInput = tblInput.withColumn(strColumn, F.array_distinct(F.flatten(strColumn)))
        return tblInput

    ##################################################
    #####                                        #####
    #####          [Data Prep Functions]         #####
    #####    These are functions for data prep;  #####
    #####     presented in order of execution    #####
    #####                                        #####
    ##################################################
    # tested and vetted
    def format_column_names(self, tblInput,strTableName):
        """
        #Inputs:
        1. [tblInput]: Target table.
        2. [strTableName]: Table name.

        #Process/Output
        1. Rename columns to [strTableName] + [column name] so that joining tables wont cause ambiguity problems.
        """
        return tblInput.select([F.col(strColumnName).alias(f"{strTableName}_{strColumnName}") for strColumnName in tblInput.columns])
    
    # tested and vetted
    def filter_to_market(self, dictblInputs, dicParams, boolVerbose = False):
        """
        #Inputs:
        1. [dictblInputs]: a dictionary of table that will be cleaned for analysis and filter for viability for master table. Format of dictionary as follows:
        - dictblInputs = {
            'tblService': tblService,
            'tblVehicle': tblVehicle,
            'tblCustomer': tblCustomer,
            }
        2. [dicParams]: Optional. Dictionary of parameters.

        #Process/Outputs::
        1. Returns a dictionary with the filtered tables based on the input dictionary of tables. This also contains metadata of before and after filters.
        2. Filtering rules:
        - **tblVehicle**: Filtered to include only rows where the 'MAKE' matches the values in `dicParams['tblVehicle_MAKE']`.
        - **tblService**: Filtered to include rows that can be merged with the filtered 'tblVehicle'.
        - **tblCustomer**: Filtered to include rows that can be merged with the filtered 'tblService'.
            (Note: The filtering of each table is dependent on the results of prior table filtering.)

        """
        strColumnName_tblVehicle_MAKE = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_MAKE')
        strColumnName_tblService_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_VEHIID')
        strColumnName_tblVehicle_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_VEHIID')
        strColumnName_tblService_CUSTNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_CUSTNO')
        strColumnName_tblCustomer_CUSTNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblCustomer_CUSTNO')
        lisstrCarMakes = self.get_parameter_value(dicParams = dicParams, strKey = 'Car_Makes')
        dicResults = {
            'tblVehicle': None,
            'tblService': None,
            'tblCustomer': None,
            'intCountBeforeVehicle':None,
            'intCountBeforeService':None,
            'intCountBeforeCustomer':None,
            'intCountAfterVehicle':None,
            'intCountAfterService':None,
            'intCountAfterCustomer':None,
        }
        
        ##################################################
        #####                                        #####
        #####       Step 1: Filter Vehicle Table     #####
        #####                                        #####
        ##################################################
        try:
            tblTemp = dictblInputs['tblVehicle']
            dicResults['intCountBeforeVehicle'] = tblTemp.count()
            dicResults['tblVehicle'] = tblTemp.filter(F.lower(F.col(strColumnName_tblVehicle_MAKE)).isin(lisstrCarMakes))
            dicResults['intCountAfterVehicle'] = dicResults['tblVehicle'].count()
            if boolVerbose:
                print("tblVehicle")
                print(f"Removed rows: {dicResults['intCountBeforeVehicle'] - dicResults['intCountAfterVehicle'] }")
                print(f"Remaining rows: {dicResults['intCountAfterVehicle'] }")
            if dicResults['intCountAfterVehicle'] == 0 :
                raise(f"[[filter_to_market()]] Error 'tblVehicle' values too few, skipping the rest of the processes.")
        except Exception as e:
            print(f"[[filter_to_market()]] Error 'tblVehicle' not defined, skipping filter for this: {e}")
        ##################################################
        #####                                        #####
        #####       Step 2: Filter Service Table     #####
        #####                                        #####
        ##################################################
        try:
            tblTemp = dictblInputs['tblService']
            dicResults['intCountBeforeService'] = tblTemp.count()
            lisstrValidVEHIIDs = [rowRow[strColumnName_tblVehicle_VEHIID] for rowRow in dicResults['tblVehicle'].select(strColumnName_tblVehicle_VEHIID).distinct().collect()]
            dicResults['tblService'] = tblTemp.filter(F.lower(F.col(strColumnName_tblService_VEHIID)).isin(lisstrValidVEHIIDs))
            dicResults['intCountAfterService'] = dicResults['tblService'].count()
            if boolVerbose:
                print("tblService")
                print(f"Removed rows: {dicResults['intCountBeforeService'] - dicResults['intCountAfterService'] }")
                print(f"Remaining rows: {dicResults['intCountAfterService'] }")
        except Exception as e:
            print(f"[[filter_to_market()]] Error 'tblService' not defined, skipping filter for this: {e}")
        ##################################################
        #####                                        #####
        #####      Step 3: Filter Customer Table     #####
        #####                                        #####
        ##################################################
        try:
            tblTemp = dictblInputs['tblCustomer']
            dicResults['intCountBeforeCustomer'] = tblTemp.count()
            lisstrValidCUSTNOs= [rowRow[strColumnName_tblService_CUSTNO] for rowRow in dicResults['tblService'].select(strColumnName_tblService_CUSTNO).distinct().collect()]
            dicResults['tblCustomer']  = tblTemp.filter(F.lower(F.col(strColumnName_tblCustomer_CUSTNO)).isin(lisstrValidCUSTNOs))
            dicResults['intCountAfterCustomer'] = dicResults['tblCustomer'].count()
            if boolVerbose:
                print("tblCustomer")
                print(f"Removed rows: {dicResults['intCountBeforeCustomer'] - dicResults['intCountAfterCustomer'] }")
                print(f"Remaining rows: {dicResults['intCountAfterCustomer'] }")
        except Exception as e:
            print(f"[[filter_to_market()]] Error 'tblCustomer' not defined, skipping filter for this: {e}")
        ##################################################
        #####                                        #####
        #####         Step 4: Filter VIN Table       #####
        #####                                        #####
        ##################################################
        try:
            tblTemp = dictblInputs['tblVIN']
            dicResults['intCountBeforeVIN'] = tblTemp.count()
            dicResults['tblVIN'] = tblTemp.withColumn(
                'tblVIN_GRP',
                F.when(
                    (F.col('tblVIN_GRP')=='JA') |
                    (F.col('tblVIN_GRP')=='LR') |
                    (F.col('tblVIN_GRP')=='RR') , 
                    'JLR'
                ).otherwise(F.col('tblVIN_GRP'))
            ).filter(
                F.col('tblVIN_GRP').isin(
                    'BW',
                    'JLR',
                    'UD'
                )
            )
            
            dicResults['intCountAfterVIN'] = dicResults['tblVIN'].count()
            if boolVerbose:
                print("tblVIN")
                print(f"Removed rows: {dicResults['intCountBeforeVIN'] - dicResults['intCountAfterVIN'] }")
                print(f"Remaining rows: {dicResults['intCountAfterVIN'] }")
        except Exception as e:
            print(f"[[filter_to_market()]] Error 'tblVIN' not defined, skipping filter for this: {e}")
        return dicResults

    # tested and vetted
    ## notable impact:
    ### 1. filters model year to just 1980 and beyond
    def fix_vehicles(self, tblInput, dicParams=None):
        """
        #Inputs:
        1. [tblInput]: Target table.
        2. [dicParams]: Optional. Dictionary of parameters.
        #Process/Output:
        1. Formats values in tblVehicle_MAKE, as there are different naming formats for 1 value.
        2. Fixes disguised null values in tblVehicle_MODEL and tblVehicle_MCODE and tblVehicle_MYEAR.
        3. Fixes erroneous values in tblVehicle_MYEAR as not are year values.
        4. Adds new column 'new_Make_Model` which concatenates tblVehicle_MAKE and tblVehicle_MODEL for better ML purposes.
        """
        def fix_myear(strValue):
            if strValue:
                try:
                    strValue += 0
                    strValue = int(strValue)
                except:
                    try:
                        strValue = int(strValue[-4:])
                    except:
                        strValue = None
                return strValue
            else:
                return None

        udf_fix_MYEAR = F.udf(fix_myear, IntegerType())
        strColumnName_tblVehicle_MAKE = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_MAKE')
        strColumnName_tblVehicle_MODEL = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_MODEL')
        strColumnName_tblVehicle_MCODE = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_MCODE')
        strColumnName_tblVehicle_MYEAR = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_MYEAR')
        strColumnName_tblVehicle_PURCHPR = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_PURCHPR')
        strColumnName_new_Make_Model = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Make_Model')

        ##################################################
        #####                                        #####
        #####      Step 1: Format values of MAKE     #####
        #####               MODEL, MCODE,            #####
        #####                                        #####
        ##################################################
        tblResult = tblInput.withColumn(
            strColumnName_tblVehicle_MAKE,
            F.lower(F.col(strColumnName_tblVehicle_MAKE))
        ).withColumn(
            strColumnName_tblVehicle_MODEL,
            F.lower(F.col(strColumnName_tblVehicle_MODEL))
        ).withColumn(
            strColumnName_tblVehicle_MCODE,
            F.lower(F.col(strColumnName_tblVehicle_MCODE))
        )
        tblResult = tblResult.withColumn(
            strColumnName_tblVehicle_MAKE,
            F.when((F.col(strColumnName_tblVehicle_MAKE)=='range rover') |
                    (F.col(strColumnName_tblVehicle_MAKE)=='rover') , 'land rover').otherwise(F.col(strColumnName_tblVehicle_MAKE))
        )
        ##################################################
        #####                                        #####
        #####         Step 2: Disguised Nulls        #####
        #####                                        #####
        ##################################################
        tblResult = tblResult.withColumn(
            strColumnName_tblVehicle_MODEL,
            F.when(
                (F.col(strColumnName_tblVehicle_MODEL) == '') |
                (F.col(strColumnName_tblVehicle_MODEL) == ' ') , None
            ).otherwise(F.col(strColumnName_tblVehicle_MODEL))
        ).withColumn(
            strColumnName_tblVehicle_MCODE,
            F.when(
                (F.col(strColumnName_tblVehicle_MCODE) == '') |
                (F.col(strColumnName_tblVehicle_MCODE) == ' ') , None).otherwise(F.col(strColumnName_tblVehicle_MCODE))
        )
        ##################################################
        #####                                        #####
        #####           Step 3: Fix MYEAR            #####
        #####                                        #####
        ##################################################
        tblResult = tblResult.withColumn(
            strColumnName_tblVehicle_MYEAR,
            udf_fix_MYEAR(F.col(strColumnName_tblVehicle_MYEAR))
        )
        ##################################################
        #####                                        #####
        #####         Step 4: new_Make_Model         #####
        #####                                        #####
        ##################################################
        tblResult = tblResult.withColumn(
        strColumnName_new_Make_Model,
            F.concat_ws(
                "__", 
                strColumnName_tblVehicle_MAKE,
                strColumnName_tblVehicle_MODEL
            )
        )
        ##################################################
        #####                                        #####
        #####     Step 5: impute PURCHPR & MYEAR     #####
        #####                                        #####
        ##################################################
        # we impute by getting the avg value in a group, the reason we do is this is because cars have different prices despite same model code and year due to different dealers, discounts, optional add ons, seasonal pricing.
        winSpecA = W.Window.partitionBy(strColumnName_new_Make_Model,strColumnName_tblVehicle_MYEAR)
        tblResult = tblResult.withColumn(
            strColumnName_tblVehicle_PURCHPR,
            F.coalesce(
                F.col(strColumnName_tblVehicle_PURCHPR),
                F.avg(strColumnName_tblVehicle_PURCHPR).over(winSpecA)
            )
        )
        return tblResult

    # tested and vetted
    def drop_duplicates(self, tblInput, dicParams = None, boolVerbose = False):
        """
        #Inputs:
        1. [tblInput]: Target table.
        2. [lisstrColumnNamesTarget]: Optional. List of column names to consider for duplicate removal. If none, all columns are considered.

        #Process/Output
        1. Drops duplicate rows from the input table based on the specified list of column names (if any).
        2. If a column in `lisstrColumnNamesTarget` does not exist in the input table, it will be ignored.
        3. Can report back information how many rows were removed.
        4. Can report back invalid column name targets.
        """
        lisstrColumnNamesTarget = self.get_parameter_value(dicParams = dicParams, strKey = 'lisstrColumnNames')
        if lisstrColumnNamesTarget:
            setColumnNamesTableInput = set(tblInput.columns)
            setColumnNamesTarget = set(lisstrColumnNamesTarget)
            lisstrValidColumnNames = list(setColumnNamesTableInput & setColumnNamesTarget)
            if len(lisstrValidColumnNames)>0:
                tblResult = tblInput.dropDuplicates(subset=lisstrValidColumnNames)
            lisstrInvalidColumnNames = list(setColumnNamesTarget - setColumnNamesTableInput)
            if boolVerbose & len(lisstrInvalidColumnNames)>0:
                print(f'Invalid column names: {lisstrInvalidColumnNames}')
        else:
            tblResult = tblInput.dropDuplicates()
        if boolVerbose:
            intOriginalCount = tblInput.count() 
            intResultCount = tblResult.count()
            print(f'Removed rows: {intOriginalCount - intResultCount}')
            print(f'Remaining rows: {intResultCount}')
        return tblResult
    
    # tested and vetted
    def drop_nulls(self, tblInput, lisstrColumnNamesTarget = None, boolVerbose = False):
        """
        #Inputs:
        1. [tblInput]: Target table.
        2. [lisstrColumnNamesTarget]: Optional. List of column names to consider for null removal. If none, primary key columns are used instead.

        #Process/Output
        1. Drops null rows from the input table based on the specified list of column names (if any).
        2. If a column in `lisstrColumnNamesTarget` does not exist in the input table, it will be ignored.
        3. Can report back information how many rows were removed.
        4. Can report back invalid column name targets.
        """
        if lisstrColumnNamesTarget:
            setColumnNamesTableInput = set(tblInput.columns)
            setColumnNamesTarget = set(lisstrColumnNamesTarget)
            lisstrValidColumnNames = list(setColumnNamesTableInput & setColumnNamesTarget)
            if len(lisstrValidColumnNames)>0:
                tblResult = tblInput.dropna(subset=lisstrValidColumnNames)
            lisstrInvalidColumnNames = list(setColumnNamesTarget - setColumnNamesTableInput)
            if boolVerbose & len(lisstrInvalidColumnNames)>0:
                print(f'Invalid column names: {lisstrInvalidColumnNames}')
        else:
            lisstrColumnNamesPrimaryKeys = self. get_parameter_value(dicParams = dicParams, strKey = 'Primary_Keys')
            tblResult = tblInput.dropna(subset = lisstrColumnNamesPrimaryKeys)
        if boolVerbose:
            intOriginalCount = tblInput.count() 
            intResultCount = tblResult.count()
            print(f'Removed rows: {intOriginalCount - intResultCount}')
            print(f'Remaining rows: {intResultCount}')
        return tblResult
    
    # tested and vetted
    def remove_pdi(self, tblService, tblVehicle, dicParams = None, boolVerbose = None):
        """
        #Inputs
        1. [tblService]: Target Service History table.
        2. [tblVehicle]: Auxillary Vehicle table.
        3. [dicParams]: Optional. Dictionary of parameters.
        #Process/Outputs
        1. Filter rows where the first registration date is at more than 1 day before the service date.
        2. Also removes rows that has blank first registration date.
        """
        strColumnName_tblService_SERVD = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_SERVD')
        strColumnName_tblVehicle_FREGD = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_FREGD')
        strColumnName_tblService_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_VEHIID')
        strColumnName_tblVehicle_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_VEHIID')
        tblVehicle = tblVehicle.select(
            strColumnName_tblVehicle_VEHIID,
            strColumnName_tblVehicle_FREGD
        )
        tblResult = tblService.join(
            tblVehicle,
            on = (F.col(strColumnName_tblService_VEHIID) == F.col(strColumnName_tblVehicle_VEHIID)),
            how = 'left'
        ).filter(
            F.col(strColumnName_tblVehicle_FREGD).isNotNull()
        ).filter(
            # the difference of dates should never be negative as it indicates the FREGD is more recent than SERVD, ergo marking it as PDI
            F.datediff(F.col(strColumnName_tblService_SERVD),F.col(strColumnName_tblVehicle_FREGD)) > 0
        ).drop(
            strColumnName_tblVehicle_VEHIID,
            strColumnName_tblVehicle_FREGD
        )
        if boolVerbose: 
            intOriginalCount = tblService.count()
            intResultCount = tblResult.count()
            print(f'Removed rows: {intOriginalCount - intResultCount}')
            print(f'Remaining rows: {intResultCount}')

        return tblResult
    
    # tested and vetted
    def remove_corporate_only_cars(self,tblService, tblCustomer, dicParams = None, boolVerbose = None):
        """
        #Inputs
        1. [tblService]: Target Service History table.
        2. [tblCustomer]: Auxillary Customer table.
        3. [dicParams]: Optional. Dictionary of parameters.
        #Process/Outputs
        1. Remove vehicle ids where it just contains solely either of the following:
            a. 'AN'
            b. 'V'
            c. 'VBMA'
            d. 'VBML'
        2. Keep vehicle id that even if they contain the above, as long as they have a record of 'A'. (Do note some vehicle ids can be owned by both 'A' and 'AN')
        """
        # when in doubt test this; this sample contains:
        ## 1. vehiid that are both AN and A
        ## 2. vehiid that doesnt have A 
        
        strColumnName_tblService_CUSTNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_CUSTNO')
        strColumnName_tblCustomer_CUSTNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblCustomer_CUSTNO')
        strColumnName_tblService_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_VEHIID')
        strColumnName_tblCustomer_CGRPID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblCustomer_CGRPID')
        lisstrExcludedCustomerGroup = self.get_parameter_value(dicParams = dicParams, strKey = 'lisstrExcludedCustomerGroup')
        lisstrIncludedCustomerGroup = self.get_parameter_value(dicParams = dicParams, strKey = 'lisstrIncludedCustomerGroup')

        windowSpec = W.Window.partitionBy(strColumnName_tblService_VEHIID)
        tblCustomer = tblCustomer.select(
            strColumnName_tblCustomer_CUSTNO,
            strColumnName_tblCustomer_CGRPID
        )
        tblResult = tblService.join(
            tblCustomer,
            on = (F.col(strColumnName_tblService_CUSTNO) == F.col(strColumnName_tblCustomer_CUSTNO)),
            how = 'left'
        ).withColumn(
            'Contains_AN_V_VBMA_VBML',
            F.when(F.col(strColumnName_tblCustomer_CGRPID).isin(lisstrExcludedCustomerGroup), 1).otherwise(0)
        ).withColumn(
            'Contains_A',
            F.when(F.col(strColumnName_tblCustomer_CGRPID).isin(lisstrIncludedCustomerGroup), 1).otherwise(0)
        ).withColumn(
            'Group_Contains_AN_V_VBMA_VBML',
            F.max('Contains_AN_V_VBMA_VBML').over(windowSpec)
        ).withColumn(
            'Group_Contains_A',
            F.max('Contains_A').over(windowSpec)
        ).filter(
            # Keep tblService_VEHIID groups where Group_Contains_AN_V_VBMA_VBML = 0 
            # OR if Group_Contains_AN_V_VBMA_VBML = 1, then Group_Contains_A must also be 1
            (F.col('Group_Contains_AN_V_VBMA_VBML') == 0) | 
            ((F.col('Group_Contains_AN_V_VBMA_VBML') == 1) & (F.col('Group_Contains_A') == 1))
        ).drop(
            'Contains_AN_V_VBMA_VBML', 
            'Contains_A', 
            'Group_Contains_AN_V_VBMA_VBML', 
            'Group_Contains_A',
            strColumnName_tblCustomer_CUSTNO,
            strColumnName_tblCustomer_CGRPID
        )
        
        if boolVerbose: 
            intOriginalCount = tblService.count() 
            intResultCount = tblResult.count()
            print(f'Removed rows: {intOriginalCount - intResultCount}')
            print(f'Remaining rows: {intResultCount}')

        return tblResult

    # tested and vetted
    def join_gsal_and_grow(self, tblInputGROW, tblInputGSAL, dicParams = None,  boolVerbose=False):
        """
        #Inputs:
        1. [tblInputGROW]: Work Order Row table.
        2. [tblInputGSAL]: Work Order Sales table.
        3. [dicParams]: Optional. Dictionary of parameters.
        4. [boolVerbose]: Optional. False by deault. Shows process results at each step. If you want to see, enter a valid gsalid. 

        #Process/Output
        1. Returns a table containing master table of work order talbes.
        """
        strColumnName_tblGROW_GSALID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_GSALID')
        strColumnName_tblGROW__UNITID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW__UNITID')
        strColumnName_tblGROW_DISCPC = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_DISCPC')
        strColumnName_tblGROW_ITEM = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_ITEM')
        strColumnName_tblGROW_ITEMNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_ITEMNO')
        strColumnName_tblGROW_NAME = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_NAME')
        strColumnName_tblGROW_GROWID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_GROWID') 
        strColumnName_tblGROW_is_discounted = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_is_discounted')
        strColumnName_tblGSAL_GSALID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGSAL_GSALID')
        strColumnName_tblGSAL__UNITID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGSAL__UNITID')
        strColumnName_tblGSAL_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGSAL_VEHIID')
        strColumnName_tblGSAL_WRKORDNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGSAL_WRKORDNO')

        ##################################################
        #####                                        #####
        #####   Step 1: GSAL By GSALID, SITE, VEHI   #####
        #####                                        #####
        ##################################################
        tblGSALAgg = tblInputGSAL.groupBy(
            strColumnName_tblGSAL_GSALID, 
            strColumnName_tblGSAL__UNITID, 
            strColumnName_tblGSAL_VEHIID
        ).agg(
            F.min(strColumnName_tblGSAL_WRKORDNO).alias(strColumnName_tblGSAL_WRKORDNO) # Okay to use first for aggregation because there's always just one WRKORDNO for the keys
        )
        ##################################################
        #####                                        #####
        #####      Step 2: GROW By GSALID, SITE      #####
        #####                                        #####
        ##################################################
        tblGROWAgg = tblInputGROW.groupBy(
            strColumnName_tblGROW_GSALID, 
            strColumnName_tblGROW__UNITID
        ).agg(
            F.max(F.expr(f"{strColumnName_tblGROW_DISCPC} > 0")).alias(strColumnName_tblGROW_is_discounted),
            F.collect_set(strColumnName_tblGROW_ITEM).alias(strColumnName_tblGROW_ITEM),
            F.collect_set(strColumnName_tblGROW_ITEMNO).alias(strColumnName_tblGROW_ITEMNO),
            F.collect_set(strColumnName_tblGROW_NAME).alias(strColumnName_tblGROW_NAME),
            F.collect_set(strColumnName_tblGROW_GROWID).alias(strColumnName_tblGROW_GROWID),
        )
        ##################################################
        #####                                        #####
        #####      Step 3: Join By GSALID, SITE      #####
        #####                                        #####
        ##################################################
        tblResult = tblGSALAgg.join(
            tblGROWAgg, 
            on = (
                (F.col(strColumnName_tblGSAL_GSALID) == F.col(strColumnName_tblGROW_GSALID)) &
                (F.col(strColumnName_tblGSAL__UNITID) == F.col(strColumnName_tblGROW__UNITID))
            ), 
            how = "inner"
        )            
        ##################################################
        #####                                        #####
        #####      Step 4: Agg By WO, SITE, VEHI     #####
        #####                                        #####
        ##################################################
        tblResult = tblResult.groupBy(
            # Another aggregation before next join
            # take note that a single workorder could have multiple gsalids
            strColumnName_tblGSAL_WRKORDNO, 
            strColumnName_tblGSAL__UNITID, 
            strColumnName_tblGSAL_VEHIID
        ).agg(
            F.collect_set(strColumnName_tblGSAL_GSALID).alias(strColumnName_tblGSAL_GSALID),
            F.max(strColumnName_tblGROW_is_discounted).alias(strColumnName_tblGROW_is_discounted),
            F.collect_list(strColumnName_tblGROW_ITEM).alias(strColumnName_tblGROW_ITEM),
            F.collect_list(strColumnName_tblGROW_ITEMNO).alias(strColumnName_tblGROW_ITEMNO),
            F.collect_list(strColumnName_tblGROW_NAME).alias(strColumnName_tblGROW_NAME),
            F.collect_list(strColumnName_tblGROW_GROWID).alias(strColumnName_tblGROW_GROWID),
        )
        tblResult = self.flatten_get_distinct(tblInput = tblResult, lisstrColumnTargets = [strColumnName_tblGROW_ITEM, 
                                                                                            strColumnName_tblGROW_ITEMNO, 
                                                                                            strColumnName_tblGROW_NAME, 
                                                                                            strColumnName_tblGROW_GROWID])
        return tblResult

    # tested and vetted
    ## notable impact:
    ### 1. we have a new column to refer by servd
    ### 2. we have a new column to refer by tsum
    def join_service_history_and_gsal_grow(self, tblService, tblCustomer, tblGSALGROW, dicParams = None,  boolVerbose = None):
        """
        #Inputs:
        1. [tblService]: Service History table.
        2. [tblCustomer]: Customer table.
        3. [tblGSALGROW]: Combined GSAL and GROW table.
        4. [dicParams]: Optional. Dictionary of parameters.

        #Process/Output
        1. Returns a table containing master table of work order.
        2. Also distinguish how much customer, warranty/insurance, others is paying.

        #Notable Changes Since Mikho's suggestion:
        1. add customer number
        2. add vehicle id
        3. add service id
        4. add percentage how much specific customer type is paying
        """
        lisintExcludedMileageValues = self.get_parameter_value(dicParams = dicParams, strKey = 'lisintExcludedMileageValues')

        strColumnName_tblService_SEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_SEHIID')
        strColumnName_tblService_CUSTNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_CUSTNO')
        strColumnName_tblService_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_VEHIID')
        strColumnName_tblService_WRKORDNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_WRKORDNO')
        strColumnName_tblService_UNITID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_UNITID')
        strColumnName_tblService_BILLD = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_BILLD')
        strColumnName_tblService_TSUM = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_TSUM')
        strColumnName_tblService_SERVD = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_SERVD')
        strColumnName_tblService_DISTDRIV = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_DISTDRIV')
        strColumnName_new_SERVD_Start = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Start')
        strColumnName_new_SERVD_End = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_End')
        strColumnName_new_BILLD_Start = self.get_parameter_value(dicParams = dicParams, strKey = 'new_BILLD_Start')
        strColumnName_new_BILLD_End = self.get_parameter_value(dicParams = dicParams, strKey = 'new_BILLD_End')
        strColumnName_new_VEHIID_Collected = self.get_parameter_value(dicParams = dicParams, strKey = 'new_VEHIID_Collected')
        strColumnName_new_CUSTNO_Collected = self.get_parameter_value(dicParams = dicParams, strKey = 'new_CUSTNO_Collected')
        strColumnName_new_SEHIID_Collected = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SEHIID_Collected')
        strColumnName_new_TSUM_Total = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Total')
        
        strColumnName_tblCustomer_CUSTNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblCustomer_CUSTNO')
        strColumnName_tblCustomer_CGRPID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblCustomer_CGRPID')
        strColumnName_new_CGRPID_Collected = self.get_parameter_value(dicParams = dicParams, strKey = 'new_CGRPID_Collected')
        lisstrExcludedCustomerGroup = self.get_parameter_value(dicParams = dicParams, strKey = 'lisstrExcludedCustomerGroup')
        lisstrIncludedCustomerGroup = self.get_parameter_value(dicParams = dicParams, strKey = 'lisstrIncludedCustomerGroup')
        strColumnName_new_TSUM_Coverage_Personal = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Personal')
        strColumnName_new_TSUM_Coverage_Warranty = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Warranty')
        strColumnName_new_TSUM_Coverage_Insurance = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Insurance')
        strColumnName_new_TSUM_Coverage_Others = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Others')

        strColumnName_tblGROW_ITEM = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_ITEM')
        strColumnName_tblGROW_ITEMNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_ITEMNO')
        strColumnName_tblGROW_NAME = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_NAME')
        strColumnName_tblGROW_GROWID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_GROWID') 
        strColumnName_tblGROW_is_discounted = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_is_discounted')

        strColumnName_tblGSAL_GSALID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGSAL_GSALID')
        strColumnName_tblGSAL__UNITID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGSAL__UNITID')
        strColumnName_tblGSAL_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGSAL_VEHIID')
        strColumnName_tblGSAL_WRKORDNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGSAL_WRKORDNO')
        
        ########################################
        #####                              #####
        #####   Step 1: Join SEHI & CUST   #####
        #####                              #####
        ########################################
        # identify who pays for which
        tblCustomer = tblCustomer.select(
            strColumnName_tblCustomer_CUSTNO,
            strColumnName_tblCustomer_CGRPID
        )
        tblResult = tblService.join(
            tblCustomer,
            on = (F.col(strColumnName_tblService_CUSTNO) == F.col(strColumnName_tblCustomer_CUSTNO)),
            how = 'left'
        )
        ########################################
        #####                              #####
        #####  Step 2: Agg By Work Order   #####
        #####                              #####
        ########################################
        # setting up service history to be joinable for gsal and grow
        tblResult = tblResult.groupBy(
            strColumnName_tblService_WRKORDNO, 
            strColumnName_tblService_UNITID, 
            strColumnName_tblService_VEHIID
        ).agg(
            F.sum(F.col(strColumnName_tblService_TSUM)).alias(strColumnName_new_TSUM_Total),
            F.sum(
                F.when(
                    F.col(strColumnName_tblCustomer_CGRPID).isin(lisstrIncludedCustomerGroup), 
                    F.col(strColumnName_tblService_TSUM)
                ).otherwise(0)
            ).alias(strColumnName_new_TSUM_Coverage_Personal),  # Customer payment
            F.sum(
                F.when(
                    F.col(strColumnName_tblCustomer_CGRPID) == "G", 
                    F.col(strColumnName_tblService_TSUM)
                ).otherwise(0)
            ).alias(strColumnName_new_TSUM_Coverage_Warranty),  # Warranty payment
            F.sum(  
                F.when(
                    F.col(strColumnName_tblCustomer_CGRPID) == "D", 
                    F.col(strColumnName_tblService_TSUM)
                ).otherwise(0)
            ).alias(strColumnName_new_TSUM_Coverage_Insurance),  # Insurance payment
            F.sum(
                F.when(
                    F.col(strColumnName_tblCustomer_CGRPID).isin(lisstrExcludedCustomerGroup), 
                    F.col(strColumnName_tblService_TSUM)
                ).otherwise(0)
            ).alias(strColumnName_new_TSUM_Coverage_Others),  # Other payments
            F.min(strColumnName_tblService_SERVD).alias(strColumnName_tblService_SERVD), 
            F.min(strColumnName_tblService_DISTDRIV).alias(strColumnName_tblService_DISTDRIV),
            F.min(strColumnName_tblService_BILLD).alias(strColumnName_tblService_BILLD),
            F.collect_set(strColumnName_tblService_VEHIID).alias(strColumnName_new_VEHIID_Collected),  # Keep all vehicle IDs, because we need to connect vehicle table with this
            F.collect_set(strColumnName_tblService_CUSTNO).alias(strColumnName_new_CUSTNO_Collected),  # Keep all customer numbers, because we need to connect customer table with this
            F.collect_set(strColumnName_tblService_SEHIID).alias(strColumnName_new_SEHIID_Collected),  # Keep all service id, just in case future join is needed
            F.collect_set(strColumnName_tblCustomer_CGRPID).alias(strColumnName_new_CGRPID_Collected)
        )
        ########################################
        #####                              #####
        ##### Step 3: Join SEHI GSAL GROW  #####
        #####                              #####
        ########################################
        # join sehi and gsal and grow
        tblResult = tblResult.join(
                tblGSALGROW, 
                on = ((F.col(strColumnName_tblService_WRKORDNO) == F.col(strColumnName_tblGSAL_WRKORDNO)) &
                        (F.col(strColumnName_tblService_UNITID) == F.col(strColumnName_tblGSAL__UNITID)) &
                        (F.col(strColumnName_tblService_VEHIID) == F.col(strColumnName_tblGSAL_VEHIID))), 
                how = "left"
        )
        ########################################
        #####                              #####
        #####    Step 4: Agg By Mileage    #####
        #####                              #####
        ########################################
        # get again important details but treating excluded mileage as separate workorder: excluded mileages includes 1 & 0
        tblResultExcluded = tblResult.filter( 
            # treat the service SERVD independently; treat the distdriv indepedently 
            F.col(strColumnName_tblService_DISTDRIV).isin(lisintExcludedMileageValues) 
        ).withColumn(
            strColumnName_new_SERVD_Start, 
            F.col(strColumnName_tblService_SERVD)
        ).withColumn(
            strColumnName_new_SERVD_End, 
            F.col(strColumnName_tblService_SERVD)
        ).withColumn(
            strColumnName_new_BILLD_Start, 
            F.col(strColumnName_tblService_BILLD)
        ).withColumn(
            strColumnName_new_BILLD_End, 
            F.col(strColumnName_tblService_BILLD)
        ).drop(
            # removing the following because these wont exist anymore
            strColumnName_tblService_SERVD,
            strColumnName_tblService_BILLD, 
            strColumnName_tblGSAL_VEHIID, 
            strColumnName_tblGSAL__UNITID, 
            strColumnName_tblGSAL_WRKORDNO, 
            strColumnName_tblService_BILLD,
            strColumnName_tblCustomer_CUSTNO,
            strColumnName_tblCustomer_CGRPID
        )
        tblResult = tblResult.filter(
            ~F.col(strColumnName_tblService_DISTDRIV).isin(lisintExcludedMileageValues)
        ).groupBy(
            strColumnName_tblService_VEHIID,
            strColumnName_tblService_DISTDRIV
        ).agg(
            F.min(strColumnName_tblService_WRKORDNO).alias(strColumnName_tblService_WRKORDNO), # probably collect, as no more purpose beyond this point
            F.min(strColumnName_tblService_UNITID).alias(strColumnName_tblService_UNITID), # probably mode
            F.sum(strColumnName_new_TSUM_Total).alias(strColumnName_new_TSUM_Total),
            F.sum(strColumnName_new_TSUM_Coverage_Personal).alias(strColumnName_new_TSUM_Coverage_Personal),
            F.sum(strColumnName_new_TSUM_Coverage_Warranty).alias(strColumnName_new_TSUM_Coverage_Warranty),
            F.sum(strColumnName_new_TSUM_Coverage_Insurance).alias(strColumnName_new_TSUM_Coverage_Insurance),
            F.sum(strColumnName_new_TSUM_Coverage_Others).alias(strColumnName_new_TSUM_Coverage_Others),
            F.min(strColumnName_tblService_SERVD).alias(strColumnName_new_SERVD_Start),
            F.max(strColumnName_tblService_SERVD).alias(strColumnName_new_SERVD_End),
            F.min(strColumnName_tblService_BILLD).alias(strColumnName_new_BILLD_Start),
            F.max(strColumnName_tblService_BILLD).alias(strColumnName_new_BILLD_End),
            F.max(strColumnName_tblGROW_is_discounted).alias(strColumnName_tblGROW_is_discounted),
            F.collect_list(strColumnName_tblGROW_ITEM).alias(strColumnName_tblGROW_ITEM),
            F.collect_list(strColumnName_tblGROW_ITEMNO).alias(strColumnName_tblGROW_ITEMNO),
            F.collect_list(strColumnName_tblGROW_NAME).alias(strColumnName_tblGROW_NAME),
            F.collect_list(strColumnName_tblGSAL_GSALID).alias(strColumnName_tblGSAL_GSALID),
            F.collect_list(strColumnName_tblGROW_GROWID).alias(strColumnName_tblGROW_GROWID),
            F.collect_list(strColumnName_new_SEHIID_Collected).alias(strColumnName_new_SEHIID_Collected),
            F.collect_list(strColumnName_new_VEHIID_Collected).alias(strColumnName_new_VEHIID_Collected),
            F.collect_list(strColumnName_new_CUSTNO_Collected).alias(strColumnName_new_CUSTNO_Collected),
            F.collect_list(strColumnName_new_CGRPID_Collected).alias(strColumnName_new_CGRPID_Collected)
        )
        # merging exlucded mileage that were treated as separate workorder
        tblResult = self.flatten_get_distinct(
            tblInput = tblResult, 
            lisstrColumnTargets = [strColumnName_tblGROW_ITEM, 
                                    strColumnName_tblGROW_ITEMNO, 
                                    strColumnName_tblGROW_NAME, 
                                    strColumnName_tblGSAL_GSALID,
                                    strColumnName_tblGROW_GROWID,
                                    strColumnName_new_SEHIID_Collected,
                                    strColumnName_new_VEHIID_Collected,
                                    strColumnName_new_CUSTNO_Collected,
                                    strColumnName_new_CGRPID_Collected
                                    ]
        )
        tblResult =  tblResult.unionByName(tblResultExcluded)
        if boolVerbose:
            print(f'Original service table line item count: {tblService.count()}')
            print(f'Remaining line item count: {tblResult.count()}')

        return tblResult

    # tested and vetted
    def filter_to_service_maintenance(self, tblService, dicParams = None, boolVerbose = None):
        """
        #Inputs:
        1. [tblService]: Target pyspark table that is service history.
        2. [dicParams]: Optional. Dictionary of parameters.

        #Process/Outputs:
        1. Returns a PySpark table retaining only rows who are identified as serviced maintenance via if GROW name or item has 'eļļa'.
        - basically we check if `oil change` or `oil filter` value exists, indicating its a maintenance service.
        - this also assumes `tblGROW_ITEM` and `tblGROW_NAME` are array type.
        2. This obviously requires GROW and GSAL joined to SEHI.
        """
        strColumnName_tblGROW_ITEM = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_ITEM')
        strColumnName_tblGROW_NAME = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_NAME')
        
        tblResult = tblService.filter(
            (F.expr(f"exists({strColumnName_tblGROW_NAME}, x -> lower(x) like '%eļļa%')")) |
            (F.expr(f"exists({strColumnName_tblGROW_ITEM}, x -> lower(x) like '%eļļa%')"))
        )
        if boolVerbose:
            print(f'Original service table line item count: {tblService.count()}')
            print(f'Remaining line item count: {tblResult.count()}')
        return tblResult
    
    # tested and vetted  
    ## notable impact:
    ### 1. crunches sehi records where each line item represents workorder-invoice, to sehi records of service groups 
    def get_service_groupings(self, tblInput, dicParams = None, boolVerbose = None):
        """
        #Inputs:
        1. [tblInput]: Target pyspark table.
        2. [dicParams]: Optional. Dictionary of parameters.

        #Process/Outputs:
        1. Returns a PySpark table with grouping of related services (this doesnt reduce the number of rows). The rules to grouping as follows:
            a. group services that are within the first two weeks (default: 15).
        2. Also adds the following features:
            a. `new_Unique_Visit_Key`: Concatenates `tblService_VEHIID`, `tblService_UNITID`, and `new_Service_Group` with double '_' 
        """
        strColumnName_tblService_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_VEHIID')
        strColumnName_tblService_WRKORDNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_WRKORDNO')
        strColumnName_tblService_UNITID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_UNITID')
        strColumnName_new_TSUM_Total = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Total')
        strColumnName_new_TSUM_Total_Previous = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Total_Previous')
        
        strColumnName_tblService_DISTDRIV = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_DISTDRIV')
        strColumnName_new_SERVD_Start = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Start')
        strColumnName_new_SERVD_End = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_End')
        strColumnName_new_BILLD_Start = self.get_parameter_value(dicParams = dicParams, strKey = 'new_BILLD_Start')
        strColumnName_new_BILLD_End = self.get_parameter_value(dicParams = dicParams, strKey = 'new_BILLD_End')
        strColumnName_new_VEHIID_Collected = self.get_parameter_value(dicParams = dicParams, strKey = 'new_VEHIID_Collected')
        strColumnName_new_CUSTNO_Collected = self.get_parameter_value(dicParams = dicParams, strKey = 'new_CUSTNO_Collected')
        strColumnName_new_SEHIID_Collected = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SEHIID_Collected')
        strColumnName_new_SERVD_Start_Difference_Previous = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Start_Difference_Previous')
        strColumnName_new_SERVD_End_Difference_Previous = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_End_Difference_Previous')
        strColumnName_new_New_Visit_Flag = self.get_parameter_value(dicParams = dicParams, strKey = 'new_New_Visit_Flag')
        strColumnName_new_Service_Group = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Service_Group')
        strColumnName_new_Unique_Visit_Key = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Unique_Visit_Key')

        strColumnName_tblGSAL_GSALID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGSAL_GSALID')
        strColumnName_tblGROW_ITEM = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_ITEM')
        strColumnName_tblGROW_ITEMNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_ITEMNO')
        strColumnName_tblGROW_NAME = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_NAME')
        strColumnName_tblGROW_GROWID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_GROWID') 
        strColumnName_tblGROW_is_discounted = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_is_discounted')
        
        strColumnName_new_CGRPID_Collected = self.get_parameter_value(dicParams = dicParams, strKey = 'new_CGRPID_Collected')
        strColumnName_new_TSUM_Coverage_Personal = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Personal')
        strColumnName_new_TSUM_Coverage_Warranty = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Warranty')
        strColumnName_new_TSUM_Coverage_Insurance = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Insurance')
        strColumnName_new_TSUM_Coverage_Others = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Others')
        
        winSpec = W.Window.partitionBy(strColumnName_tblService_VEHIID).orderBy(strColumnName_new_SERVD_Start)
        intDefault_Service_Grouping = 15

        tblResult = tblInput.withColumn(
            strColumnName_new_SERVD_Start_Difference_Previous, 
            F.abs(F.datediff(strColumnName_new_SERVD_Start, F.lag(strColumnName_new_SERVD_Start).over(winSpec)))
        ).withColumn(
            strColumnName_new_SERVD_End_Difference_Previous, 
            F.abs(F.datediff(strColumnName_new_SERVD_Start, F.lag(strColumnName_new_SERVD_End).over(winSpec)))
        ).withColumn(
            strColumnName_new_New_Visit_Flag, 
            F.expr(f"case when ({strColumnName_new_SERVD_Start_Difference_Previous} <= {intDefault_Service_Grouping}) or ({strColumnName_new_SERVD_End_Difference_Previous} <= {intDefault_Service_Grouping}) then 0 else 1 end")
        ).withColumn(
            strColumnName_new_Service_Group, 
            F.sum(strColumnName_new_New_Visit_Flag).over(winSpec)
        )

        tblResult = tblResult.groupBy(
            strColumnName_tblService_VEHIID, 
            strColumnName_new_Service_Group
        ).agg(
            F.min(strColumnName_tblService_WRKORDNO).alias(strColumnName_tblService_WRKORDNO), # probably collect, but service grouping requires groupby work order
            F.min(strColumnName_tblService_UNITID).alias(strColumnName_tblService_UNITID), # probably mode
            F.sum(strColumnName_new_TSUM_Total).alias(strColumnName_new_TSUM_Total),
            F.sum(strColumnName_new_TSUM_Coverage_Personal).alias(strColumnName_new_TSUM_Coverage_Personal),
            F.sum(strColumnName_new_TSUM_Coverage_Warranty).alias(strColumnName_new_TSUM_Coverage_Warranty),
            F.sum(strColumnName_new_TSUM_Coverage_Insurance).alias(strColumnName_new_TSUM_Coverage_Insurance),
            F.sum(strColumnName_new_TSUM_Coverage_Others).alias(strColumnName_new_TSUM_Coverage_Others),
            F.min(strColumnName_new_SERVD_Start).alias(strColumnName_new_SERVD_Start),
            F.max(strColumnName_new_SERVD_End).alias(strColumnName_new_SERVD_End),
            F.min(strColumnName_new_BILLD_Start).alias(strColumnName_new_BILLD_Start),
            F.max(strColumnName_new_BILLD_End).alias(strColumnName_new_BILLD_End),
            F.max(strColumnName_tblService_DISTDRIV).alias(strColumnName_tblService_DISTDRIV),
            F.max(strColumnName_tblGROW_is_discounted).alias(strColumnName_tblGROW_is_discounted),
            F.collect_list(strColumnName_tblGROW_ITEM).alias(strColumnName_tblGROW_ITEM),
            F.collect_list(strColumnName_tblGROW_ITEMNO).alias(strColumnName_tblGROW_ITEMNO),
            F.collect_list(strColumnName_tblGROW_NAME).alias(strColumnName_tblGROW_NAME),
            F.collect_list(strColumnName_tblGSAL_GSALID).alias(strColumnName_tblGSAL_GSALID),
            F.collect_list(strColumnName_tblGROW_GROWID).alias(strColumnName_tblGROW_GROWID),
            F.collect_list(strColumnName_new_SEHIID_Collected).alias(strColumnName_new_SEHIID_Collected),
            F.collect_list(strColumnName_new_VEHIID_Collected).alias(strColumnName_new_VEHIID_Collected),
            F.collect_list(strColumnName_new_CUSTNO_Collected).alias(strColumnName_new_CUSTNO_Collected),
            F.collect_list(strColumnName_new_CGRPID_Collected).alias(strColumnName_new_CGRPID_Collected)
        ).withColumn(
            strColumnName_new_TSUM_Total_Previous,
            F.lag(strColumnName_new_TSUM_Total).over(winSpec)
        ).withColumn(
            strColumnName_new_Unique_Visit_Key, 
            F.concat_ws(
                "__", 
                strColumnName_tblService_VEHIID, 
                strColumnName_tblService_UNITID, 
                strColumnName_new_Service_Group
            )
        )
        tblResult = self.flatten_get_distinct(tblInput = tblResult, lisstrColumnTargets = [strColumnName_tblGROW_ITEM, 
                                                                                            strColumnName_tblGROW_ITEMNO, 
                                                                                            strColumnName_tblGROW_NAME, 
                                                                                            strColumnName_tblGSAL_GSALID,
                                                                                            strColumnName_tblGROW_GROWID,
                                                                                            strColumnName_new_SEHIID_Collected,
                                                                                            strColumnName_new_VEHIID_Collected,
                                                                                            strColumnName_new_CUSTNO_Collected,
                                                                                            strColumnName_new_CGRPID_Collected
                                                                                            ])
        if boolVerbose:
            print(f'Original service table line item count: {tblInput.count()}')
            print(f'Remaining line item count: {tblResult.count()}')
        return tblResult

    # tested and vetted
    def join_service_history_and_vehicle(self, tblService, tblVehicle, dicParams = None, boolVerbose = None):
        """
        #Inputs:
        1. [tblInput]: Target pyspark table.
        2. [dicParams]: Optional. Dictionary of parameters.

        #Process/Outputs:
        1. Boring way to join the table and record it.
        """
        strColumnName_tblService_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_VEHIID')
        strColumnName_tblVehicle_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_VEHIID')
        tblResult = tblService.join(
            tblVehicle,
            on = (F.col(strColumnName_tblService_VEHIID) == (F.col(strColumnName_tblVehicle_VEHIID))),
            how = 'inner'
        )
        if boolVerbose:
            print(f'Original service table line item count: {tblService.count()}')
            print(f'Final line item count: {tblResult.count()}')
        return tblResult
    
    # tested and vetted
    def get_vins(self, tblInput, tblInputVINS, tblInputCustomer, dicParams = None, boolVerbose = False):
        """
        # Inputs
            1. [tblInput]: pyspark table. Master table preferably, that has vehiid where the vins would be attached to.
            2. [tblInputVINS]: pyspark table. VIN table that contains the vinsid.
            3. [tblInputCustomer]; pyspark table. Customer table that gives insight wether the VIN is owned by customer or other type.
            4. [dicParams]: Optional. Dictionary of parameters.
        # Process/Outputs
            1. Attaches latest vin id values to the input table where vin is also owned by a customer type.
        """
        
        strColumnName_tblVehicle_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_VEHIID')
        strColumnName_tblVehicle_HCUSTNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_HCUSTNO')
        strColumnName_tblCustomer_CGRPID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblCustomer_CGRPID')
        strColumnName_tblCustomer_CUSTNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblCustomer_CUSTNO')
        strColumnName_tblVIN_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVIN_VEHIID')
        strColumnName_tblVIN_BBILLD = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVIN_BBILLD')
        strColumnName_tblVIN_CLOSED = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVIN_CLOSED')
        strColumnName_tblVIN_VINSID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVIN_VINSID')
        strColumnName_tblVIN_HCUSTNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVIN_HCUSTNO')
        ##################################################
        #####                                        #####
        #####         Step 1: Get Latest VIN         #####
        #####                                        #####
        ##################################################

        # Select only latest VIN id; because again VINS shows the VEHI and its VINS (VINS changes overtime, and changes are recorded under ORIG_VINSID)
        winSpec = W.Window.partitionBy(strColumnName_tblVIN_VEHIID).orderBy(F.desc(strColumnName_tblVIN_BBILLD),F.desc(strColumnName_tblVIN_CLOSED),F.desc(strColumnName_tblVIN_VINSID))
        tblInputVINS = tblInputVINS.withColumn(
            'temp_rank',
            F.rank().over(winSpec)
        ).filter(
            F.col('temp_rank')==1
        ).drop('temp_rank')
        
        ##################################################
        #####                                        #####
        #####      Step 2: Select Customer Only      #####
        #####                  VIN                   #####
        #####                                        #####
        ##################################################
        tblInputVINS = tblInputVINS.join(
            tblInputCustomer.select(strColumnName_tblCustomer_CUSTNO,strColumnName_tblCustomer_CGRPID),
            on = [F.col(strColumnName_tblVIN_HCUSTNO) == F.col(strColumnName_tblCustomer_CUSTNO)],
            how = 'inner'
        ).withColumn(
            'tblVIN_HCUSTNO_CGRPID',
            F.col(strColumnName_tblCustomer_CGRPID)
        ).drop(
            strColumnName_tblCustomer_CUSTNO,
            strColumnName_tblCustomer_CGRPID
        )

        ##################################################
        #####                                        #####
        #####           Step 3: Join Table           #####
        #####                                        #####
        ##################################################

        tblResult = tblInput.join(
            tblInputVINS.select(strColumnName_tblVIN_VEHIID,strColumnName_tblVIN_HCUSTNO,strColumnName_tblVIN_VINSID),
            on=[
                F.col(strColumnName_tblVehicle_VEHIID) == F.col(strColumnName_tblVIN_VEHIID),
                F.col(strColumnName_tblVehicle_HCUSTNO) == F.col(strColumnName_tblVIN_HCUSTNO)
            ],
            how = 'left' # not inner join, because not all cars came from latvia, can come from another country neighboring latvia
        )
        if boolVerbose:
            print('count of initial: ',tblInput.count())
            print('count of result: ',tblResult.count())
        
        ##################################################
        #####                                        #####
        #####          Step 4: Return Table          #####
        #####                                        #####
        ##################################################
        return tblResult
       
    # tested and vetted
    def create_initial_warranty_details(self, tblInput, dicParams = None):
        """
        #Inputs:
        1. [tblInput]: Target pyspark table.
        2. [dicParams]: Optional. Dictionary of parameters.

        #Process/Outputs:
        1. Returns a pyspark table adding inital warranty details set by market, the new columns are as follows:
        - **new_Warranty_Period_In_Days**: 3 years (1095 days) or 5 years (1825 days).
        - **new_Warranty_Period_By_Mileage**: 100,000 or 150,000 mileage.
        """
        lisstrCarMakes = self.get_parameter_value(dicParams = dicParams, strKey = 'Car_Makes')
        dtDateWarrantyPolicyChange = F.to_date(
            F.lit(self.get_parameter_value(dicParams=dicParams, strKey='Date_Warranty_Policy_Change')),
            "yyyy-MM-dd'T'HH:mm:ss'Z'"
        )
        strColumnName_tblVehicle_FREGD = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_FREGD')
        strColumnName_tblVehicle_MAKE = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_MAKE')
        strColumnName_new_Warranty_Period_In_Days = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Warranty_Period_In_Days')
        strColumnName_new_Warranty_Period_By_Mileage = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Warranty_Period_By_Mileage')
        tblResult = tblInput.withColumn(
            # determines if warranty period if 3 years or 5 years, depending if car is registered before 1st july 2022
            strColumnName_new_Warranty_Period_In_Days,
            F.when((F.datediff(dtDateWarrantyPolicyChange,strColumnName_tblVehicle_FREGD)<=0) & (F.col(strColumnName_tblVehicle_MAKE).isin(lisstrCarMakes)), 1825).otherwise(1095)
        ).withColumn(
            strColumnName_new_Warranty_Period_By_Mileage,
            F.when((F.datediff(dtDateWarrantyPolicyChange,strColumnName_tblVehicle_FREGD)<=0) & (F.col(strColumnName_tblVehicle_MAKE).isin(lisstrCarMakes)), 150000).otherwise(100000)
        )
        return tblResult

    # tested and vetted
    def create_feature_related_to_age(self, tblInput, dicParams = None):
        """
        #Inputs:
        1. [tblInput]: Target pyspark table. This requires tsum 
        2. [dicParams]: Optional. Dictionary of parameters.

        #Process/Outputs:
        1. Returns a pyspark table containing features about the warranties, this adds the following features:
            - **new_Age_Of_Car**: Integer. The age of car in days.
            - **new_Expected_Visits_By_DISTDRIV**: Integer. The number of vists the car should have made based on mileage.
            - **new_Expected_Vs_Actual_Visits_By_DISTDRIV**: Float. Out of expected visits by mileage, how many visits did this car had.
            - **new_Expected_Visits_By_Age_Of_Car**: Integer. The number of vists the car should have made based on car age.
            - **new_Expected_Vs_Actual_Visits_By_Age_Of_Car**: Float. Out of expected visits by car age, how many visits did this car had.
            - **new_Days_Passed_Previous_Service**: Integer. The number of days passed since previous visit.
            - **new_Days_Passed_First_Service**: Integer. The number of days passed since first visit.
        """
        strColumnName_tblService_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_VEHIID') 
        strColumnName_new_SERVD_Start = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Start') 
        strColumnName_new_SERVD_End = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_End')
        strColumnName_tblVehicle_FREGD = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_FREGD')
        strColumnName_tblService_DISTDRIV = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_DISTDRIV')
        strColumnName_new_Service_Group = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Service_Group')
        strColumnName_new_Age_Of_Car = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Age_Of_Car') 
        strColumnName_new_Age_Of_Car_Previous = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Age_Of_Car_Previous') 
        strColumnName_new_Expected_Visits_By_DISTDRIV = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Expected_Visits_By_DISTDRIV')
        strColumnName_new_Expected_Vs_Actual_Visits_By_DISTDRIV = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Expected_Vs_Actual_Visits_By_DISTDRIV')
        strColumnName_new_Expected_Visits_By_Age_Of_Car = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Expected_Visits_By_Age_Of_Car')
        strColumnName_new_Expected_Vs_Actual_Visits_By_Age_Of_Car = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Expected_Vs_Actual_Visits_By_Age_Of_Car')
        strColumnName_new_Days_Passed_Previous_Service = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Days_Passed_Previous_Service')
        strColumnName_new_Mileage_Passed_Previous_Service = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Mileage_Passed_Previous_Service')
        strColumnName_new_Days_Passed_First_Service = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Days_Passed_First_Service')
        intServiceWindowCarAge = self.get_parameter_value(dicParams = dicParams, strKey = 'intServiceWindowCarAge') #int(tblInitialMaster.agg(F.avg('Days_From_Start_Previous_Service')).first()[0]) # average value is 115.60120829645592 days or roughly 4 months
        intServiceWindowMileage = self.get_parameter_value(dicParams = dicParams, strKey = 'intServiceWindowMileage') #10000

        windowSpecB = W.Window.partitionBy(
            strColumnName_tblService_VEHIID,
        ).orderBy(
            F.asc(strColumnName_new_SERVD_Start),
        )
        tblResult = tblInput.withColumn(
            # Step 1: age of car
            # table VEHI column FREGD ; this is first registration date
            strColumnName_new_Age_Of_Car, 
            F.datediff(F.col(strColumnName_new_SERVD_Start),F.col(strColumnName_tblVehicle_FREGD))
        ).withColumn(
            # Step 2: age of car from previous service
            strColumnName_new_Age_Of_Car_Previous,
            F.lag(strColumnName_new_Age_Of_Car).over(windowSpecB)
        ).withColumn(
            # Step 3: expected visits by mileage
            strColumnName_new_Expected_Visits_By_DISTDRIV,
            F.floor(F.col(strColumnName_tblService_DISTDRIV)/intServiceWindowMileage)
        ).withColumn(
            # Step 4: actual_visits / expected_visits_by_mileage
            strColumnName_new_Expected_Vs_Actual_Visits_By_DISTDRIV,
            F.col(strColumnName_new_Service_Group) / F.col(strColumnName_new_Expected_Visits_By_DISTDRIV)
        ).withColumn(
            # Step 5: expected visits by car age
            strColumnName_new_Expected_Visits_By_Age_Of_Car,
            F.floor(F.col(strColumnName_new_Age_Of_Car)/intServiceWindowCarAge)
        ).withColumn(
            # Step 6: actual_visits / expected_visits_by_car_age
            strColumnName_new_Expected_Vs_Actual_Visits_By_Age_Of_Car,
            F.col(strColumnName_new_Service_Group)/ F.col(strColumnName_new_Expected_Visits_By_Age_Of_Car)
        ).withColumn(
            # Step 7: Days After Previous Service
            strColumnName_new_Days_Passed_Previous_Service,
            F.when(
                F.lag(strColumnName_new_SERVD_Start).over(windowSpecB).isNull(), -1  # Set -1 if no lag value exists
            ).otherwise(
                F.datediff(F.col(strColumnName_new_SERVD_Start), F.lag(strColumnName_new_SERVD_End).over(windowSpecB))  # Regular calculation
            )
        ).withColumn(
            # Step 8: Mileages After Previous Service
            strColumnName_new_Mileage_Passed_Previous_Service,
            F.when(
                F.lag(strColumnName_new_SERVD_Start).over(windowSpecB).isNull(), -1  # Set -1 if no lag value exists
            ).otherwise(
                F.col(strColumnName_tblService_DISTDRIV) - F.lag(strColumnName_tblService_DISTDRIV).over(windowSpecB)  # Regular calculation
            )
        ).withColumn(
            # Step 9: Days After First Service
            strColumnName_new_Days_Passed_First_Service,
            F.when(
                F.col(strColumnName_new_SERVD_Start) == F.first(strColumnName_new_SERVD_Start).over(windowSpecB), -1  # Set -1 if theres no other service exists
            ).otherwise(
                F.datediff(F.col(strColumnName_new_SERVD_Start), F.first(strColumnName_new_SERVD_End).over(windowSpecB))  # Regular calculation
            )
        )
        return tblResult
    
    # tested and vetted
    def create_feature_related_to_warranty(self, tblInput, dicParams = None):
        """
        #Inputs:
        1. [tblInput]: Target pyspark table. This requires tsum 
        2. [dicParams]: Optional. Dictionary of parameters.

        #Process/Outputs:
        1. Returns a pyspark table containing features about the warranties, this adds the following features:
            - **new_Warranty_Period_Left_In_Days**: Integer. The number of days left before warranty expires.
            - **new_Warranty_Period_By_Mileage**: Integer. The number of mileage left before warranty expires.
            - **new_Warranty_Is_Expired**: Integer 1 or 0. 1 Indicates if new_Warranty_Period_Left_In_Days is negative value or if new_Warranty_Period_By_Mileage is negative value, indicating the warranty has expired already.
            - **new_TSUM_Coverage_Warranty_Percentage**: Float. How much percent new_TSUM_Coverage_Warranty_Percentage covered in new_TSUM_Total. 
            - **new_TSUM_Coverage_Insurance_Percentage**: Float. How much percent new_TSUM_Coverage_Insurance_Percentage covered in new_TSUM_Total. 
            - **new_TSUM_Coverage_Combined_Insurance_Warranty**: Float. How much percent new_TSUM_Coverage_Warranty_Percentage & new_TSUM_Coverage_Insurance_Percentage covered in new_TSUM_Total.
        2. This requires the `new_Age_Of_Car` to be computed beforehand.
        """
        
        #strColumnName_new_SERVD_Start = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Start')
        #strColumnName_tblVehicle_FREGD = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_FREGD')
        strColumnName_tblService_DISTDRIV = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_DISTDRIV')
        strColumnName_new_Warranty_Period_In_Days = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Warranty_Period_In_Days')
        strColumnName_new_Age_Of_Car = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Age_Of_Car') 
        strColumnName_new_Warranty_Period_By_Mileage = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Warranty_Period_By_Mileage')
        strColumnName_new_Warranty_Period_Left_In_Days = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Warranty_Period_Left_In_Days')
        strColumnName_new_Warranty_Period_Left_By_Mileage = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Warranty_Period_Left_By_Mileage')
        strColumnName_new_Warranty_Is_Expired = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Warranty_Is_Expired')
        strColumnName_new_Warranty_Will_Expire_Next_Service_By_Days = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Warranty_Will_Expire_Next_Service_By_Days')
        strColumnName_new_Warranty_Will_Expire_Next_Service_By_Mileage = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Warranty_Will_Expire_Next_Service_By_Mileage')
        strColumnName_new_TSUM_Total = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Total')
        strColumnName_new_TSUM_Coverage_Warranty = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Warranty')
        strColumnName_new_TSUM_Coverage_Insurance = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Insurance')
        strColumnName_new_TSUM_Coverage_Warranty_Percentage = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Warranty_Percentage')
        strColumnName_new_TSUM_Coverage_Insurance_Percentage = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Insurance_Percentage')
        strColumnName_new_TSUM_Coverage_Combined_Insurance_Warranty = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Combined_Insurance_Warranty')
        
        intServiceWindowCarAge = self.get_parameter_value(dicParams = dicParams, strKey = 'intServiceWindowCarAge')
        intServiceWindowMileage = self.get_parameter_value(dicParams = dicParams, strKey = 'intServiceWindowMileage')
        
        tblResult = tblInput.withColumn(
            strColumnName_new_Warranty_Period_Left_In_Days,
            F.col(strColumnName_new_Warranty_Period_In_Days) - F.col(strColumnName_new_Age_Of_Car)
        ).withColumn(
            strColumnName_new_Warranty_Period_Left_By_Mileage,
            F.col(strColumnName_new_Warranty_Period_By_Mileage) - F.col(strColumnName_tblService_DISTDRIV)
        ).withColumn(
            strColumnName_new_Warranty_Is_Expired,
            F.when(
                (F.col(strColumnName_new_Warranty_Period_Left_In_Days) < 0) |
                (F.col(strColumnName_new_Warranty_Period_Left_By_Mileage) < 0),
                1 
            ).otherwise(0)
        ).withColumn(
            strColumnName_new_Warranty_Will_Expire_Next_Service_By_Days,
            F.when(F.col(strColumnName_new_Warranty_Period_Left_In_Days) < F.lit(intServiceWindowCarAge), 1).otherwise(0)
        ).withColumn(
            strColumnName_new_Warranty_Will_Expire_Next_Service_By_Mileage,
            F.when(F.col(strColumnName_new_Warranty_Period_Left_By_Mileage) < F.lit(intServiceWindowMileage), 1).otherwise(0)
        ).withColumn(
            strColumnName_new_TSUM_Coverage_Warranty_Percentage,
            F.col(strColumnName_new_TSUM_Coverage_Warranty) / F.col(strColumnName_new_TSUM_Total)
        ).withColumn(
            strColumnName_new_TSUM_Coverage_Insurance_Percentage,
            F.col(strColumnName_new_TSUM_Coverage_Insurance) / F.col(strColumnName_new_TSUM_Total)
        ).withColumn(
            strColumnName_new_TSUM_Coverage_Combined_Insurance_Warranty,
            F.col(strColumnName_new_TSUM_Coverage_Insurance_Percentage) + F.col(strColumnName_new_TSUM_Coverage_Warranty_Percentage)
        ).fillna({
            strColumnName_new_TSUM_Coverage_Warranty_Percentage: 0.0,
            strColumnName_new_TSUM_Coverage_Insurance_Percentage: 0.0,
            strColumnName_new_TSUM_Coverage_Combined_Insurance_Warranty: 0.0
        })
        return tblResult

    # tested and vetted
    def create_feature_related_to_economy(self, tblInput, dicParams = None):
        """
        #Inputs:
        1. [tblInput]: Target pyspark table. This requires tsum 
        2. [dicParams]: Optional. Dictionary of parameters.

        #Process/Outputs:
        1. Returns a pyspark table containing features about the economy, this adds the following features:
            - **new_Economy_2008_Financial_Crisis**: Integer 1 or 0. Indicates if service is within this year or next year of the event.
            - **new_Economy_2020_Covid_19**: Integer 1 or 0. Indicates if service is within this year or next year of the event.
            - **new_Economy_2022_Ukraine_Russia_War**: Integer 1 or 0. Indicates if service is within this year or next year of the event.
        """
        strColumnName_tblService_SERVD = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Start')
        strColumnName_new_Economy_2008_Financial_Crisis = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Economy_2008_Financial_Crisis')
        strColumnName_new_Economy_2020_Covid_19 = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Economy_2020_Covid_19')
        strColumnName_new_Economy_2022_Ukraine_Russia_War = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Economy_2022_Ukraine_Russia_War')
        tblInput = tblInput.withColumn(
            strColumnName_new_Economy_2008_Financial_Crisis,
            F.when( (F.year(strColumnName_tblService_SERVD) >= 2008) & 
                    (F.year(strColumnName_tblService_SERVD) <= 2009) ,
                    (F.year(strColumnName_tblService_SERVD) - 2008) + 1).otherwise(0)
        ).withColumn(
            strColumnName_new_Economy_2020_Covid_19,
            F.when( (F.year(strColumnName_tblService_SERVD) >= 2020) & 
                    (F.year(strColumnName_tblService_SERVD) <= 2021) ,
                    (F.year(strColumnName_tblService_SERVD) - 2020) + 1).otherwise(0)
        ).withColumn(
            strColumnName_new_Economy_2022_Ukraine_Russia_War,
            F.when( (F.year(strColumnName_tblService_SERVD) >= 2022) & 
                    (F.year(strColumnName_tblService_SERVD) <= 2023) ,
                    (F.year(strColumnName_tblService_SERVD) - 2022) + 1).otherwise(0)
        )
        return tblInput

    # tested and vetted
    def create_feature_related_to_work_order(self, tblInput, dicParams = None):
        """
        #Inputs:
        1. [tblInput]: Target pyspark table. This requires tsum 
        2. [dicParams]: Optional. Dictionary of parameters.

        #Process/Outputs:
        1. Returns a pyspark table containing features about the work details, this adds the following features:
            - **new_GROWID_Count**: Integer. The count of works done, based on the collected size of growid.
            - **new_Service_Duration**: Integer. The duration of the service; between identified service date start and service date end.
        """
        strColumnName_tblService_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_VEHIID')
        strColumnName_new_SERVD_Start = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Start')
        strColumnName_new_SERVD_End = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_End')
        strColumnName_new_Service_Duration = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Service_Duration')
        strColumnName_new_Service_Duration_Previous = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Service_Duration_Previous')
        strColumnName_tblGROW_GROWID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblGROW_GROWID')
        strColumnName_new_GROWID_Count = self.get_parameter_value(dicParams = dicParams, strKey = 'new_GROWID_Count')
        
        winSpec  = W.Window.partitionBy(strColumnName_tblService_VEHIID).orderBy(strColumnName_new_SERVD_Start)
        tblResult = tblInput.withColumn(
            strColumnName_new_GROWID_Count, 
            F.when(F.col(strColumnName_tblGROW_GROWID).isNull(), 0).otherwise(F.size(F.col(strColumnName_tblGROW_GROWID)))
        ).withColumn(
            strColumnName_new_Service_Duration,
            F.datediff(
                F.col(strColumnName_new_SERVD_End),
                F.col(strColumnName_new_SERVD_Start),
            ) + F.lit(1) # this addidition is necessary, because if first and last service are same, originally the result is 0, which doesnt make sense because it should count as 1 day in service
        ).withColumn(
            strColumnName_new_Service_Duration_Previous,
            F.col(strColumnName_new_Service_Duration) - F.lag(strColumnName_new_Service_Duration).over(winSpec)
        )
        return tblResult

    # tested and vetted
    def create_feature_target_variables(self, tblInput, dicParams = None, boolVerbose = False):
        """
        # Inputs
        1. [tblInput]: pyspark table. Target input table to be modified.
        2. [dicParams]: Optional. Dictionary of parameters.
        # Process/Outputs
        1. Creates the following columns for the input dataframe:
            -**new_SERVD_Next**: date. The service date
            -**new_SERVD_Next_Days**: int. The number of days for next service date.
            -**new_Churn_Flag**: bool. 1 indicates it churned when the next service days is over the number defined in `intServiceWindowCarAge`, otherwise 0.
        """
        strColumnName_tblService_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_VEHIID')
        strColumnName_tblService_SERVD = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Start')
        strColumnName_new_SERVD_Next = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Next')
        strColumnName_new_SERVD_Next_Days = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Next_Days')
        strColumnName_new_Churn_Flag = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Churn_Flag')
        strColumnName_new_Churn_Flag_Previous = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Churn_Flag_Previous')
        intServiceWindowCarAge = self.get_parameter_value(dicParams = dicParams, strKey = 'intServiceWindowCarAge')

        windowSpecA = W.Window.partitionBy(
            strColumnName_tblService_VEHIID,
        ).orderBy(
            F.desc(strColumnName_tblService_SERVD),
        )
        
        tblResult = tblInput.withColumn(
            strColumnName_new_SERVD_Next,
            F.lag(strColumnName_tblService_SERVD,1).over(windowSpecA)
        ).withColumn(
            strColumnName_new_SERVD_Next_Days, 
            F.datediff(F.col(strColumnName_new_SERVD_Next),F.col(strColumnName_tblService_SERVD))
        ).withColumn(
            strColumnName_new_Churn_Flag,
            F.when(
                F.col(strColumnName_new_SERVD_Next_Days) > intServiceWindowCarAge,
                1
            ).when(
                (F.col(strColumnName_new_SERVD_Next).isNull()) & 
                (F.datediff(F.current_date(), F.col(strColumnName_tblService_SERVD)) > intServiceWindowCarAge),
                1
            ).when(
                F.col(strColumnName_new_SERVD_Next_Days) < intServiceWindowCarAge,
                0
            ).otherwise(None)
        ).withColumn(
            strColumnName_new_Churn_Flag_Previous,
            F.lead(strColumnName_new_Churn_Flag).over(windowSpecA)
        )

        if boolVerbose:
            print(f'Original service table line item count: {tblInput.count()}')
            print(f'Remaining line item count: {tblResult.count()}')
        
        return tblResult
    
        # check here
    
    # tested and vetted
    def create_feature_rolling_stats(self, tblInput, dicParams = None):
        """
        #Inputs:
        1. [tblInput]: Target pyspark table.
        2. [dicParams]: Optional. Dictionary of parameters.
        #Process/Outputs:
        1. Returns a dataframe with new features consisting of rolling statistics (average, sum, and max) features over a window defined by `intLagPreviouses`. 
        2. Features calculated are in `lisstrColumnNamesPreviouses` defined in `dicParams`.
        """
        # Raw
        strColumnName_tblService_SERVD = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Start')
        strColumnName_tblService_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_VEHIID')
        # Work Order
        strColumnName_new_Service_Duration = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Service_Duration')
        strColumnName_new_GROWID_Count = self.get_parameter_value(dicParams = dicParams, strKey = 'new_GROWID_Count')
        # TSUMS
        strColumnName_new_TSUM_Coverage_Personal = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Personal')
        strColumnName_new_TSUM_Coverage_Warranty = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Warranty')
        strColumnName_new_TSUM_Coverage_Insurance = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Insurance')
        strColumnName_new_TSUM_Coverage_Others = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Coverage_Others')
        strColumnName_new_TSUM_Total = self.get_parameter_value(dicParams = dicParams, strKey = 'new_TSUM_Total')
        # Age Of Car
        strColumnName_new_Days_Passed_Previous_Service  = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Days_Passed_Previous_Service')
        strColumnName_new_Mileage_Passed_Previous_Service = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Mileage_Passed_Previous_Service')
        # Churn
        strColumnName_new_Churn_Flag = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Churn_Flag')
        intLagPreviouses = self.get_parameter_value(dicParams = dicParams, strKey = 'intLagPreviouses')

        dicstrColumnNames = {
            'Rolling_Average':[
                strColumnName_new_Days_Passed_Previous_Service,
                strColumnName_new_Mileage_Passed_Previous_Service,
                strColumnName_new_GROWID_Count,
                strColumnName_new_Service_Duration,
                strColumnName_new_TSUM_Coverage_Personal,
                strColumnName_new_TSUM_Coverage_Warranty,
                strColumnName_new_TSUM_Coverage_Insurance,
                strColumnName_new_TSUM_Coverage_Others,
                strColumnName_new_TSUM_Total,
            ],
            'Rolling_Sum':[
                strColumnName_new_Churn_Flag,
                strColumnName_new_TSUM_Coverage_Personal,
                strColumnName_new_TSUM_Coverage_Warranty,
                strColumnName_new_TSUM_Coverage_Insurance,
                strColumnName_new_TSUM_Coverage_Others,
                strColumnName_new_TSUM_Total,
            ],
            'Rolling_Max':[
                strColumnName_new_Churn_Flag,
                strColumnName_new_Days_Passed_Previous_Service,
                strColumnName_new_Mileage_Passed_Previous_Service,
                strColumnName_new_Service_Duration,
                strColumnName_new_TSUM_Coverage_Personal,
                strColumnName_new_TSUM_Coverage_Warranty,
                strColumnName_new_TSUM_Coverage_Insurance,
                strColumnName_new_TSUM_Coverage_Others,
                strColumnName_new_TSUM_Total,
            ],
            'Difference_From_Rolling_Average':[
                strColumnName_new_Days_Passed_Previous_Service,
                strColumnName_new_Mileage_Passed_Previous_Service,
                strColumnName_new_GROWID_Count,
                strColumnName_new_Service_Duration,
                strColumnName_new_TSUM_Coverage_Personal,
                strColumnName_new_TSUM_Coverage_Warranty,
                strColumnName_new_TSUM_Coverage_Insurance,
                strColumnName_new_TSUM_Coverage_Others,
                strColumnName_new_TSUM_Total,
            ],
            'Difference_From_Rolling_Max':[
                strColumnName_new_Days_Passed_Previous_Service,
                strColumnName_new_Mileage_Passed_Previous_Service,
                strColumnName_new_Service_Duration,
                strColumnName_new_TSUM_Coverage_Personal,
                strColumnName_new_TSUM_Coverage_Warranty,
                strColumnName_new_TSUM_Coverage_Insurance,
                strColumnName_new_TSUM_Coverage_Others,
                strColumnName_new_TSUM_Total,
            ]
        }

        ##################################################
        #####                                        #####
        #####      Step 1: Determine the Window      #####
        #####                                        #####
        ##################################################
        if intLagPreviouses < 1:
            print('[create_feature_previouses()] WARNING!!! `intLagPreviouses` is set to less than 1, this means that user intends to start the rolling statistics since day 1 of the service...')
            windowSpecA = W.Window.partitionBy(
                strColumnName_tblService_VEHIID,
            ).orderBy(
                F.asc(strColumnName_tblService_SERVD),
            ).rowsBetween(
                W.Window.unboundedPreceding,  # Start from the first row of the partition
                -1  # End before current row
            )
        else:
            windowSpecA = W.Window.partitionBy(
                strColumnName_tblService_VEHIID,
            ).orderBy(
                F.asc(strColumnName_tblService_SERVD),
            ).rowsBetween(
                -(intLagPreviouses), 
                -1
            )
        ##################################################
        #####                                        #####
        #####      Step 2: Calculate Roll Stats      #####
        #####                                        #####
        ##################################################
        # Step 1: Calculate the average
        for strColumnName in dicstrColumnNames['Rolling_Average']:
            tblInput = tblInput.withColumn(
                f"{strColumnName}_Rolling_{intLagPreviouses}_Service_Average",
                F.avg(strColumnName).over(windowSpecA)
            )
        # Step 2: Calculate the sum
        for strColumnName in dicstrColumnNames['Rolling_Sum']:
            tblInput = tblInput.withColumn(
                f"{strColumnName}_Rolling_{intLagPreviouses}_Service_Sum",
                F.sum(strColumnName).over(windowSpecA)
            )
        # Step 3: Calculate the max
        for strColumnName in dicstrColumnNames['Rolling_Max']:
            tblInput = tblInput.withColumn(
                f"{strColumnName}_Rolling_{intLagPreviouses}_Service_Max",
                F.max(strColumnName).over(windowSpecA)
            )
        # Step 4: Calculate the difference of current vs rolling average
        for strColumnName in dicstrColumnNames['Difference_From_Rolling_Average']:
            tblInput = tblInput.withColumn(
                f"{strColumnName}_Difference_From_Rolling_{intLagPreviouses}_Service_Average",
                F.col(strColumnName) - F.col(f"{strColumnName}_Rolling_{intLagPreviouses}_Service_Average")
                )
        # Step 5: Calculate the difference of current vs rolling max
        for strColumnName in dicstrColumnNames['Difference_From_Rolling_Max']:
            tblInput = tblInput.withColumn(
                f"{strColumnName}_Difference_From_Rolling_{intLagPreviouses}_Service_Max",
                F.col(strColumnName) - F.col(f"{strColumnName}_Rolling_{intLagPreviouses}_Service_Max")
                )
        return tblInput

    # tested and vetted
    def get_modelling_dataset(self, tblInput, dicParams = None):
        """
        # Inputs
            1. [tblInput]: pyspark table. Target input table to be modified.
            2. [dicParams]: Optional. Dictionary of parameters.
        # Process/Outputs
            1. Removes rows where churn flag cannot be identified: due to the fact that their service dates is not yet more than 365 days from the date today.
            2. Reduces columns to the top features.
        """    
        strColumnName_new_Churn_Flag = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Churn_Flag')
        lisstrModellingFeatures = self.get_parameter_value(dicParams = dicParams, strKey = 'lisstrModellingFeatures')        
        tblResult = tblInput.filter(
            F.col(strColumnName_new_Churn_Flag).isNotNull()
        ).select(
            *lisstrModellingFeatures
        )
        return tblResult

    ##################################################
    #####                                        #####
    #####            [Scoring Functions]         #####
    #####     These are functions for scoring;   #####
    #####     presented in order of execution    #####
    #####                                        #####
    ##################################################

    # tested and vetted
    '''def get_prediction_report_old(self, tblInput, dicParams = None, boolTest = False, boolVerbose = None):
        """
        # Inputs
            1. [tblInput]: pyspark table. Target input table to be modified.
            2. [dicParams]: Optional. Dictionary of parameters.
        # Process/Outputs
            1. Returns a table with the predicted churn flag and churn probability on those churn flags that are not identified. 
                1. This means this removes rows where churn flag has been identified.
            2. This means this triggers the datarobot model to predict.
            3. This also reduces the features to the only ones specified in `tblFeatureList`
            4. Adds the following features to the resulting table:
                1. top 3 reasons for churning
                2. churn flag
                3. churn probability
                4. next expected service date
                5. VIN
        """
        strColumnName_new_Churn_Flag = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Churn_Flag')
        strColumnName_new_SERVD_Start = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Start')
        strColumnName_new_Next_Exp_Service = self.get_parameter_value(dicParams = dicParams, strKey = 'Next_Exp_Service')
        strColumnName_new_Next_Exp_Service_Days_Left = self.get_parameter_value(dicParams = dicParams, strKey = 'Next_Exp_Service_Days_Left')
        strColumnName_tblVehicle_MAKE = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_MAKE')
        strColumnName_new_Unique_Visit_Key = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Unique_Visit_Key')
        strColumnName_tblVIN_VINSID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVIN_VINSID')

        strDR_PROJECT_ID = self.get_parameter_value(dicParams = dicParams, strKey = 'DR_PROJECT_ID') 
        strDR_MODEL_ID = self.get_parameter_value(dicParams = dicParams, strKey = 'DR_MODEL_ID')
        intServiceWindowCarAge = self.get_parameter_value(dicParams = dicParams, strKey = 'intServiceWindowCarAge')
        lisstrMarketName = self.get_parameter_value(dicParams = dicParams, strKey = 'lisstrMarketName')
        self.auth_datarobot(dicParams=dicParams)

        ##################################################
        #####                                        #####
        #####      Step 1: Get Table To Predict      #####
        #####                                        #####
        ##################################################
        if boolTest:
            tblResult = tblInput.filter(
                F.col(strColumnName_new_Churn_Flag).isNull()
            ).limit(
                10
            )
        else:
            tblResult = tblInput.filter(
                F.col(strColumnName_new_Churn_Flag).isNull()
            )

        tblResult = tblResult.withColumn(
            strColumnName_new_Next_Exp_Service,
            F.date_add(F.col(strColumnName_new_SERVD_Start),intServiceWindowCarAge)
        ).withColumn(
            strColumnName_new_Next_Exp_Service_Days_Left,
            F.coalesce(F.date_diff(F.col(strColumnName_new_Next_Exp_Service),F.current_date()), F.lit(0))
        ).filter(
            F.col(strColumnName_tblVehicle_MAKE).isin(lisstrMarketName)
        ).toPandas()
        
        ##################################################
        #####                                        #####
        #####         Step 2: Predict Proper         #####
        #####                                        #####
        ##################################################
        print("Prediction running. . .")
        # Load the trained model
        objDRModel = dr.Model.get(project=strDR_PROJECT_ID, model_id=strDR_MODEL_ID)
        # Start job and wait
        objJob = objDRModel.request_predictions(
            file_path=tblResult,
            explanation_algorithm="shap",
            max_explanations=5
        )
        tblResultPrediction = objJob.get_result_when_complete() # pandas data type

        ##################################################
        #####                                        #####
        #####         Step 3: Format Results         #####
        #####                                        #####
        ##################################################
        # tblResultPrediction will return a primary_key-less table, thus we dont know which prediction is to which vehicle id
        # thus our option is to join/concat them index wise; this is safe to do since the result is pandas dataframe just like the input table, thus the order is preserved.
        print("Formating results. . .")
        tblResult = pd.concat(
                [tblResult[[
                    strColumnName_new_Unique_Visit_Key,
                    strColumnName_tblVehicle_MAKE,
                    strColumnName_tblVIN_VINSID,
                    strColumnName_new_Next_Exp_Service,
                    strColumnName_new_Next_Exp_Service_Days_Left
                ]],
                tblResultPrediction],
                axis=1) # join by column
        tblResult = spark.createDataFrame(tblResult)
        
        # attach the prediction to the input table
        tblResult = tblResult.select(
            strColumnName_new_Unique_Visit_Key,
            strColumnName_new_Next_Exp_Service,
            strColumnName_new_Next_Exp_Service_Days_Left,
            "`class_1.0`",
            'prediction',
            'Explanation_1_feature_name',
            'Explanation_2_feature_name',
            'Explanation_3_feature_name',
            'Explanation_4_feature_name',
            'Explanation_5_feature_name'
        ).join(
            # this is correct, tblResult initially just have nulls on churn flags, then we gave prediction to it, then just attach the input table for the other information needed
            tblInput, 
            on = strColumnName_new_Unique_Visit_Key,
            how = 'left'
        ).withColumn(
            strColumnName_new_Churn_Flag,
            F.coalesce(strColumnName_new_Churn_Flag,'prediction')
        ).withColumn(
            "Date_Predicted", 
            F.current_timestamp())
        # tblResult = self.format_prediction_report(tblResult,dicParams) # formats to layman columns and values; and join it to input table

        if boolVerbose:
            print(f'line item count of initial input: {tblInput.count()}')
            print(f'line item count of final output: {tblResult.count()}')
        return tblResult
    '''
    
    def get_prediction_report(self, tblInput, dicParams = None, boolTest = False, boolVerbose = None):
        """
        # Inputs
            1. [tblInput]: pyspark table. Target input table to be modified.
            2. [dicParams]: Optional. Dictionary of parameters.
        # Process/Outputs
            1. Returns a table with the predicted churn flag and churn probability on those churn flags that are not identified. 
                1. This means this removes rows where churn flag has been identified.
            2. This means this triggers the datarobot model to predict.
            3. This also reduces the features to the only ones specified in `tblFeatureList`
            4. Adds the following features to the resulting table:
                1. top 3 reasons for churning
                2. churn flag
                3. churn probability
                4. next expected service date
                5. VIN
        """
        strColumnName_new_Churn_Flag = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Churn_Flag')
        strColumnName_new_SERVD_Start = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Start')
        strColumnName_new_Next_Exp_Service = self.get_parameter_value(dicParams = dicParams, strKey = 'Next_Exp_Service')
        strColumnName_new_Next_Exp_Service_Days_Left = self.get_parameter_value(dicParams = dicParams, strKey = 'Next_Exp_Service_Days_Left')
        strColumnName_tblVehicle_MAKE = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_MAKE')
        strColumnName_new_Unique_Visit_Key = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Unique_Visit_Key')
        strColumnName_tblVIN_VINSID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVIN_VINSID')

        strDR_PROJECT_ID = self.get_parameter_value(dicParams = dicParams, strKey = 'DR_PROJECT_ID') 
        strDR_MODEL_ID = self.get_parameter_value(dicParams = dicParams, strKey = 'DR_MODEL_ID')
        intServiceWindowCarAge = self.get_parameter_value(dicParams = dicParams, strKey = 'intServiceWindowCarAge')
        lisstrMarketName = self.get_parameter_value(dicParams = dicParams, strKey = 'lisstrMarketName')
    
        ##################################################
        #####                                        #####
        #####      Step 1: Get Table To Predict      #####
        #####                                        #####
        ##################################################
        tblResult = tblInput.filter(
            F.col(strColumnName_new_Churn_Flag).isNull()
        ).withColumn(
            strColumnName_new_Next_Exp_Service,
            F.date_add(
                F.col(strColumnName_new_SERVD_Start),
                intServiceWindowCarAge
            )
        ).withColumn(
            strColumnName_new_Next_Exp_Service_Days_Left,
            F.coalesce(
                F.date_diff(
                    F.col(strColumnName_new_Next_Exp_Service),
                    F.current_date()
                ), 
                F.lit(0)
            )
        ).filter(
            F.col(strColumnName_tblVehicle_MAKE).isin(lisstrMarketName)
        ).toPandas()

        ##################################################
        #####                                        #####
        #####         Step 2: Predict Proper         #####
        #####                                        #####
        ##################################################

        # create if else for method to access model
        ###   --- start --- 

        ## if model is registerd
        client = mlflow.MlflowClient()
        model = mlflow.sklearn.load_model(MODEL_URI_ALIAS)
        ## if model is still in experiment
        best_run_id = "2e27e208bdcc4c06a596caa5a777f5fe"
        model_uri = f"runs:/{best_run_id}/model"
        model_version = mlflow.pyfunc.load_model(model_uri)
        
        ###   --- end --- 


        df_scored['churn_1_PREDICTION'] = model.predict_proba(df_scoring)[:,1]
df_scored['churn_PREDICTION'] = model.predict(df_scoring)

        ##################################################
        #####                                        #####
        #####           Step 3: Apply SHAP           #####
        #####                                        #####
        ##################################################

        ##################################################
        #####                                        #####
        #####         Step 4: Format Results         #####
        #####                                        #####
        ##################################################


    # tested and vetted
    def format_prediction_report(self, tblInput, dicParams = None):
        """
        too lazy to explain this one uwu
        """
        strColumnName_tblService_DISTDRIV = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_DISTDRIV')
        strColumnName_tblService_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_VEHIID') 
        strColumnName_tblVehicle_MAKE = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_MAKE')
        strColumnName_tblVIN_VINSID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVIN_VINSID')
        strColumnName_new_Unique_Visit_Key = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Unique_Visit_Key')
        strColumnName_new_Age_Of_Car = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Age_Of_Car') 
        strColumnName_new_Service_Group = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Service_Group')
        strColumnName_new_SERVD_Start = self.get_parameter_value(dicParams = dicParams, strKey = 'new_SERVD_Start')
        strColumnName_new_Next_Exp_Service = self.get_parameter_value(dicParams = dicParams, strKey = 'Next_Exp_Service')
        strColumnName_new_Next_Exp_Service_Days_Left = self.get_parameter_value(dicParams = dicParams, strKey = 'Next_Exp_Service_Days_Left')
        tblFeatureList = self.get_parameter_value(dicParams = dicParams, strKey = 'tblFeatureList')

        def map_to_categories(item_list):
            return list(set(dicMapCategories[item] for item in item_list if item in dicMapCategories))
        def rename_features(feature_list):
            return [dicMapRename.get(feature, feature) for feature in feature_list]
        ##################################################
        #####                                        #####
        #####       Step 1: Create the initial       #####
        #####              excel format              #####
        #####                                        #####
        ##################################################
        '''
        tblResult = tblInput.select( # fixxxx hereeee
            strColumnName_tblVehicle_MAKE,
            strColumnName_new_Unique_Visit_Key,
            F.col(strColumnName_tblVIN_VINSID).alias('VIN'),
            F.col(strColumnName_tblService_VEHIID).alias('Vehicle ID'), 
            F.col("`class_1.0`").alias('Churn Probability'), # Churn Probability
            F.col('prediction').alias('Prediction'), # Churn Probability
            F.col(strColumnName_new_SERVD_Start).alias('Date of last visit'), # Date of last visit
            F.col(strColumnName_new_Next_Exp_Service).alias('Next expected scheduled date'), # Next expected scheduled date
            F.col(strColumnName_new_Next_Exp_Service_Days_Left).alias('Next expected scheduled days left'),
            F.col(strColumnName_new_Service_Group).alias('Scheduled Service Sequence'), # Scheduled Service Sequence
            F.array('Explanation_1_feature_name','Explanation_2_feature_name','Explanation_3_feature_name','Explanation_4_feature_name','Explanation_5_feature_name').alias('Top 5 Feature Impact'), # Top 5 Feature Impact
            (F.col(strColumnName_tblService_DISTDRIV)/(F.col(strColumnName_new_Age_Of_Car)/F.lit(30))).alias('KM per month'), # KM per month
            *[F.col(rowRow['Technical Name']).alias(rowRow['Layman Name']) for rowRow in tblFeatureList.collect() if rowRow['Include In Model'] == 1] # edit here
        ).withColumn(
            'Prediction',
            F.when(F.col('Prediction')==1,'Churned').otherwise('Not_Churned')
        )'''

        tblResult = tblInput.withColumn(
            'KM Per Month',
            F.col(strColumnName_tblService_DISTDRIV)/(F.col(strColumnName_new_Age_Of_Car)/F.lit(30))
        ).withColumn(
            'Top 5 Feature Impact',
            F.array('Explanation_1_feature_name','Explanation_2_feature_name','Explanation_3_feature_name','Explanation_4_feature_name','Explanation_5_feature_name')
        ).select(
            *[F.col(rowRow['Technical Name']).alias(rowRow['Layman Name']) for rowRow in tblFeatureList.collect() if rowRow['Include In Model'] == 2],
            *[F.col(rowRow['Technical Name']).alias(rowRow['Layman Name']) for rowRow in tblFeatureList.collect() if rowRow['Include In Model'] == 1]
        ).withColumn(
            'Prediction',
            F.when(F.col('Prediction')==1,'Churned').otherwise('Not_Churned')
        )
        ##################################################
        #####                                        #####
        #####        Step 2: Create the top 5        #####
        #####                categories              #####
        #####                                        #####
        ##################################################
        dicMapCategories = {}
        for rowRow in tblFeatureList.collect():
            dicMapCategories.update({rowRow['Technical Name']:rowRow['Feature Category']})
        udfMapCategories = F.udf(map_to_categories, ArrayType(StringType()))
        tblResult = tblResult.withColumn("top_categories", udfMapCategories(F.col("Top 5 Feature Impact")))
        ##################################################
        #####                                        #####
        #####        Step 3: Convert all feats       #####
        #####                to layman               #####
        #####                                        #####
        ##################################################
        dicMapRename = {}
        for rowRow in tblFeatureList.collect():
            dicMapRename.update({rowRow['Technical Name']:rowRow['Layman Name']})
        for strColumn in tblResult.columns:
            if strColumn in dicMapRename.keys():
                tblResult = tblResult.withColumnRenamed(strColumn,dicMapRename.get(strColumn))
        udfRenameFeatures = F.udf(rename_features, ArrayType(StringType()))
        tblResult = tblResult.withColumn("Top 5 Feature Impact", udfRenameFeatures(F.col("Top 5 Feature Impact")))
        for col_name in tblResult.columns:
            tblResult = tblResult.withColumnRenamed(col_name, col_name.replace(" ", "_"))
        return tblResult
    
    # tested and vetted
    def create_excel_report(self, tblInput, dr_storage_accnt_name = None, dicParams = None):
        """
        # Inputs
            1. [tblInput]: pyspark table. Target input table that has already been formatted, to generate an excel report for.
            2. [dicParams]: Optional. Dictionary of parameters.
        # Process/Outputs
            1. Generates an Excel report from `tblInput` and saves it to the UC volume specified in `strPathName`.
            2. The report includes only:
                - Car makes listed in `lisstrMarketName`.
                - Customers with `Next_expected_scheduled_days_left` within `intReportWindow` days.
        """    
           
        strColumnName_tblVehicle_MAKE = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_MAKE')
        lisstrMarketName = self.get_parameter_value(dicParams = dicParams, strKey = 'lisstrMarketName')
        tblFeatureList = self.get_parameter_value(dicParams = dicParams, strKey = 'tblFeatureList')
        intDaysBetweenStart = self.get_parameter_value(dicParams = dicParams, strKey = 'intDaysBetweenStart')
        intDaysBetweenEnd = self.get_parameter_value(dicParams = dicParams, strKey = 'intDaysBetweenEnd')
        strPathName = self.get_parameter_value(dicParams = dicParams, strKey = 'strPathName')
        
        ##################################################
        #####                                        #####
        #####          Step 1: Filter Table          #####
        #####                                        #####
        ##################################################
        
        tblInput = tblInput.filter(
            F.col(strColumnName_tblVehicle_MAKE).isin(lisstrMarketName)
        ).filter(
            (F.col('Next_expected_scheduled_days_left') <= intDaysBetweenEnd) &
            (F.col('Next_expected_scheduled_days_left') >= intDaysBetweenStart)
        ).withColumn(
            'Next_expected_scheduled_date',
            F.to_timestamp(F.col('Next_expected_scheduled_date'), "yyyy-MM-dd'T'HH:mm:ss.SSSXXX")
        )
        
        ##################################################
        #####                                        #####
        #####    Step 2: Format Input Table Values   #####
        #####                                        #####
        ##################################################
        for col_name in tblInput.columns:
            tblInput = tblInput.withColumnRenamed(col_name, col_name.replace("_", " "))
        tblInput = tblInput.withColumn(
            'Top 5 Feature Impact', 
            F.concat_ws(", ", F.col('Top 5 Feature Impact'))
        )
        ##################################################
        #####                                        #####
        #####  Step 3: Format Feature Table Values   #####
        #####                                        #####
        ##################################################
        '''tblFeatureList = tblFeatureList.filter(
            F.col('Include In Model').isin([1,2])
        ).groupBy(
            'Feature Category'
        ).agg(
            F.collect_set('Layman Name').alias('Collected Layman Name')
        ).filter(
            ~F.col('Feature Category').isin(['Target','Primary Key','External'])
        ).orderBy(
            F.asc('Feature Category')
        ) 
        dicFeatureList = {
            row["Feature Category"]: row["Collected Layman Name"] for row in tblFeatureList.collect()
        }'''

        tblFeatureList = tblFeatureList.toPandas()
        tblFeatureList = tblFeatureList[
            tblFeatureList['Include In Model'].isin([1, 2])
        ].groupby('Feature Category', as_index=False).agg({
            'Layman Name': lambda x: set(x)  # Collect unique values into a set
        }).rename(columns={'Layman Name': 'Collected Layman Name'})

        tblFeatureList = tblFeatureList[
            ~tblFeatureList['Feature Category'].isin(['Target', 'Primary Key', 'External'])
        ].sort_values(by='Feature Category', ascending=True)

        '''dicFeatureList = {
            row["Feature Category"]: list(row["Collected Layman Name"]) for _, row in tblFeatureList.iterrows()
        }'''

        dicFeatureList = {
            'Basic':[
                'VIN',
                'License Number',
                'Serial Number',
                'Churn Probability',
                'Prediction',
                'Date of last visit',
                'Next expected scheduled date',
                'Next expected scheduled days left',
                'Current Mileage',
                'KM Per Month',
                'Vehicle ID',
                'Top 5 Feature Impact',
            ]
        }
        for _,row in tblFeatureList.iterrows():
            if row["Feature Category"] != 'Basic':
                dicFeatureList[row["Feature Category"]] = list(row["Collected Layman Name"])

        '''
        {'Basic': ['Serial Number', 'Vehicle ID', 'Churn Probability', 'Date of last visit', 'Prediction', 'Next expected scheduled date', 'License Number', 'Next expected scheduled days left', 'Top 5 Feature Impact', 'VIN', 'KM Per Month'], 'Car': ['Vehicle price', 'Vehicle make and model', 'Vehicle model year', 'Vehicle age during PMS', 'Current Mileage', 'Vehicle age previous PMS'], 'Finance': ['Maximum invoice value since first PMS', 'Total invoice value charged to customer during PMS', 'Total invoice value during PMS', 'Average invoice value during PMS since first PMS', 'Invoice value charged to customer deviation from average invoice value charged to customer', 'Total invoice value previous PMS', 'Cumulative sum of invoice value since first PMS'], 'Service Regularity': ['Cumulative count of late last 3 PMS', 'Maximum interval of days since last 3 PMS', 'Actual over Expected Car Age', 'Cummulative count of late visits since first PMS', 'Service inteval deviation from service interval average', 'Service duration deviation from average service duration', 'Actual over Expected Mileage', 'Mileage interval deviation from mileage interval maximum', 'Service Sequence', 'Average number of days since first PMS', 'Mileage since last PMS', 'Number of days since last PMS', 'Average duration of PMS since first PMS', 'Mileage interval deviation from mileage interval average within 3 PMS', 'Count of jobs during PMS', 'Service inteval deviation from service interval average within 3 PMS', 'Is this PMS late', 'Mileage interval deviation from mileage interval average'], 'Warranty': ['Invoice value charged to warranty deviation from average invoice value charged to warranty', 'Number of days remaining for warranty', 'Total invoice value charged to warranty during PMS', 'Mileage remaining for warranty', 'Average invoice value charged to warranty since first PMS']}
        
        '''
        '''
        dicFeatureList = {
        'Basic': [
            'tblVIN VINSID',  # Chassis Number
            'tblVehicle VEHIID',
            'Churn Probability',  # Churn Probability
            'Prediction',
            'Date of last visit',  # Date of last visit
            'Next expected scheduled date',  # Next expected scheduled date
            'Next expected scheduled days left',
            'Scheduled Service Sequence',  # Scheduled Service Sequence
            'Top 5 Feature Impact',
            'KM per month'
            ]
        }
        dicFeatureList.update({
            row["Feature Category"]: row["Collected Layman Name"] for row in tblFeatureList.collect()
        })

        '''
        
        
        ##################################################
        #####                                        #####
        #####      Step 4: Prepare Excel Report      #####
        #####                                        #####
        ##################################################
        SHEET_NAME = 'test_excel_latvia'
        df = tblInput.toPandas()
        categories = dicFeatureList
        HEADER_HEIGHT = 45
        COLUMN_WIDTH = 20
        BORDER_COLOR = "D9D9D9"
        FILL_COLORS = ['152C47', '224874', '2b5b93', '3977c0', '8DB4E2']

        def header_format(cell, FillColor, FontColor):
            cell.fill = PatternFill(fill_type="solid", fgColor=FillColor)
            cell.font = Font(name="Century Gothic", size=10, bold=True, color=FontColor)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            
        def cell_value(cell, value, cell_number_format):
            cell.value = value
            cell.border = Border(
                left  =Side(border_style=BORDER_THIN, color=BORDER_COLOR),
                right =Side(border_style=BORDER_THIN, color=BORDER_COLOR),
                top   =Side(border_style=BORDER_THIN, color=BORDER_COLOR),
                bottom=Side(border_style=BORDER_THIN, color=BORDER_COLOR))
            cell.font = Font(name="Century Gothic", size=10)
            cell.number_format = cell_number_format
        
        def num_fmt(x):
            if   x.dtype == 'object':         return '@'
            elif x.dtype == 'datetime64[ns]': return 'yyyy-mm-dd'
            else:
                average = x.abs().mean()
                if   average > 1:               return "#,#"
                elif average < 0.01:            return "#,##0.0000"
                else:                           return "#,##0.00"
        
        number_format = df.apply(num_fmt)

        start_col = 2
        start_row = 2

        wb = openpyxl.Workbook()
        wb.remove(wb['Sheet'])
        wb.create_sheet(SHEET_NAME)
        ws = wb[SHEET_NAME]
        ws.sheet_view.showGridLines = False

        # Wrap Header Rows
        ws.row_dimensions[start_row].height = HEADER_HEIGHT

        # Freeze panes
        ws.freeze_panes = "E3"

        # Loop over categories
        for (cat, cols), fcolor in zip(categories.items(), FILL_COLORS): # cat == category; cols = layman
            #data = df[cols].copy()
            if cat == "Basic":
                data = df.loc[:, cols].copy()  # Ensures order is preserved
            else:
                data = df[cols].copy()
            # Format Category Sections
            if cat != 'Basic':
                data[cat] = df['top categories'].apply(lambda x: 'YES' if cat in x else '')
                data = data[[cat] + cols] 
            # Include Column Names in DataFrame
            data = data.T.reset_index(0).T
            # Iterate through Columns
            for col_num in range(start_col, start_col + data.shape[1]):
                # Set Column Width
                ws.column_dimensions[get_column_letter(col_num)].width = COLUMN_WIDTH
                if data.iloc[0, col_num - start_col] in number_format:
                    cell_num_format = number_format[data.iloc[0, col_num - start_col]]
                else:
                    cell_num_format = '@'
                # Iterate through Rows
                for row_num in range(start_row, start_row + data.shape[0]):
                    # Add Data
                    cell = ws.cell(row=row_num, column=col_num)
                    cell_value(cell, data.iloc[row_num - start_row, col_num - start_col], cell_num_format)
                    # Format Headers
                    if row_num == start_row:
                        header_format(cell, FillColor=fcolor, FontColor='FFFFFF')
                    # Category Formatting
                    elif cat != 'Basic':
                        # Conditional Formatting if 'YES'
                        if col_num == start_col:
                            if cell.value == 'YES':
                                header_format(cell, FillColor='FFFFFF', FontColor=fcolor)
                        # If Top Feature, add color
                        if ws.cell(start_row, col_num).value in df['Top 5 Feature Impact'].iloc[row_num - start_row - 1]:
                            cell.fill = PatternFill(fill_type="solid", fgColor='FFFFC1')
        
            # Thinner column spacing (does not work after grouping)
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 3
            # Group Columns
            if cat != 'Basic':
                ws.column_dimensions.group(get_column_letter(start_col + 1),
                                        get_column_letter(start_col + data.shape[1]),
                                        hidden=True) 
            start_col += data.shape[1] + 1

        ##################################################
        #####                                        #####
        #####        Step 5: Save Excel Report       #####
        #####                                        #####
        ##################################################
        # Convert Workbook to Bytes
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)  # Move cursor to the start
        
        # Save Directly to the Volume:
        if len(lisstrMarketName) > 1:
            strMarketname = '_'.join(lisstrMarketName)
        else:
            strMarketname = lisstrMarketName[0]
        strMarketname = strMarketname.replace(" ", "_")

        timestamp = datetime.now().strftime("%Y%m%d")
        # Update the filename to include the timestamp
        with open(f"{strPathName}churn_report_{strMarketname}_{timestamp}.xlsx", "wb") as f: 
            f.write(out.getvalue())

        return True

    ##################################################
    #####                                        #####
    #####          [Modelling Functions]         #####
    #####    These are functions for modelling;  #####
    #####     presented in order of execution    #####
    #####                                        #####
    ##################################################

    def train_model(self,tblTrain:DataFrame, strPathExperiment:str,
                    strPathHistoricalExperiments:str=None, dicParams:dict=None,
                    strNameExperiment:str=None, intTimeoutMinutes:int=120):
        """
        # Inputs
            1. [tblTrain]: pyspark table. Target input table to be modified.
            2. [strPathExperiment]: string. Path to the directory in the workspace to save the generated notebooks and experiments.
            3. [strPathHistoricalExperiments]: string. Table location in unity catalog containing historical details of experiments. This table is created by this function, so this means if you don't have this, you haven't recorded any experiment.
            4. [strNameExperiment]: string. Name for the MLflow experiment that AutoML creates.
            5. [intTimeoutMinutes]: integer. Maximum time to wait for AutoML trials to complete. Longer timeouts allow AutoML to run more trials and identify a model with better accuracy.
            6. [dicParams]: Optional. Dictionary of parameters.
        # Process
            1. Takes the `tblTrain` as well as all `dicParams` contents to train multiple models with `automl.classify()`.
            2. Saves all `TrialInfo` (also known as experiment info) created by the `automl.classify()` as table in `strPathHistoricalExperiments`.
        # Outputs
            1. Returns an AutoMLSummary.best_trial object (essentially the details to the best model).
        """
        strColumnName_new_Churn_Flag = self.get_parameter_value(dicParams = dicParams, strKey = 'new_Churn_Flag')
        strEvalMetric = self.get_parameter_value(dicParams = dicParams, strKey = 'Evaluation_Metric')
        intPosLabel = self.get_parameter_value(dicParams = dicParams, strKey = 'Positive_Label')
        strEvalMetric = self.get_parameter_value(dicParams = dicParams, strKey = 'Evaluation_Metric')
        dicModelExperiments = {
                # user added
                # 'rank': [], # soon to be added, algorithm will rank
                'date_trained':[],
                # default trial info details
                'notebook_path':[],
                'notebook_url':[],
                'artifact_uri':[],
                'mlflow_run_id':[],
                'metrics':[],
                'params':[],
                'model_path':[],
                'model_description':[],
                'duration':[],
                'preprocessors':[],
                'evaluation_metric_score':[],
            }
        
        # Train
        objSummary = automl.classify(
            dataset=tblTrain,
            target_col=strColumnName_new_Churn_Flag,
            primary_metric=strEvalMetric,
            pos_label=intPosLabel,
            experiment_dir=strPathExperiment,
            experiment_name=strNameExperiment,
            timeout_minutes=intTimeoutMinutes,
            exclude_cols=None,
            exclude_frameworks=None,
            feature_store_lookups=None,
            imputers=None,
            time_col=None,
            split_col=None,
            sample_weight_col=None,
            max_trials=None,
            data_dir=None,
        )

        # Record To Pyspark
        lisobjTrials = [objTrialInfo for objTrialInfo in objSummary.trials]
        lisobjTrials.append(objSummary.best_trial) # adding this here because i speculate this isnt present in .trials()
        for objTrialInfo in lisobjTrials:
            dicModelExperiments['date_trained'].append(datetime.now())
            dicModelExperiments['notebook_path'].append(objTrialInfo.notebook_path)
            dicModelExperiments['notebook_url'].append(objTrialInfo.notebook_url)
            dicModelExperiments['artifact_uri'].append(objTrialInfo.artifact_uri)
            dicModelExperiments['mlflow_run_id'].append(objTrialInfo.mlflow_run_id)
            dicModelExperiments['metrics'].append(objTrialInfo.metrics)
            dicModelExperiments['params'].append(objTrialInfo.params)
            dicModelExperiments['model_path'].append(objTrialInfo.model_path)
            dicModelExperiments['model_description'].append(objTrialInfo.model_description)
            dicModelExperiments['duration'].append(objTrialInfo.duration)
            dicModelExperiments['preprocessors'].append(objTrialInfo.preprocessors)
            dicModelExperiments['evaluation_metric_score'].append(objTrialInfo.evaluation_metric_score)  # f1 score
        tblModelExperiments = pd.DataFrame(dicModelExperiments)
        tblModelExperiments = spark.createDataFrame(tblModelExperiments)
        tblModelExperiments.display()

        # Pull From Historical Experiment
        winSpec = W.Window.orderBy(F.desc('evaluation_metric_score'))
        if strPathHistoricalExperiments:
            boolTableExist = spark.catalog.tableExists(strPathHistoricalExperiments)
        else:
            boolTableExist = False

        lisstrColNames = tblTrain.columns
        if not boolTableExist:
            # no historical records exists 
            tblHistoricalExperiments = tblModelExperiments 
        else:
            # check if exists then proceed
            tblHistoricalExperiments = spark.table(strPathHistoricalExperiments).drop('rank')
            tblHistoricalExperiments = tblHistoricalExperiments.unionByName(
                tblModelExperiments
            )
        tblHistoricalExperiments = tblHistoricalExperiments.withColumn(
                'rank',
                F.row_number().over(winSpec)
            ).withColumn(
                'features',
                F.array(
                    [F.lit(strColName) for strColName in lisstrColNames]
                )
            )
        # Update
        tblHistoricalExperiments.write.format(
            "delta"
        ).mode(
            "overwrite"
        ).option(
            "overwriteSchema", 
            True
        ).saveAsTable(
            strPathHistoricalExperiments
        )

        return objSummary.best_trial

    ##################################################
    #####                                        #####
    #####          [Outdated Functions]          #####
    #####    These are functions for that are    #####
    #####        outdated; used by other         #####
    #####        non important notebooks         #####
    #####                                        #####
    ##################################################

    # tested and vetted: does absolutely nothing but checks missing
    def count_nulls(self, tblInput,dicParams=None):
        """
        #Inputs:
        1. [tblInput]: Target table.
        2. [dicParams]: Optional. Dictionary of parameters.
        #Process/Output:
        1. Returns a dataframe showing the following per column of dataframe:
            1. count of mising
            2. missing rate
        """
        intRowCounts = tblInput.count()
        # Step 1: Just get null count per column
        tblResult = tblInput.select(*[
            F.count(F.when(
                    F.col(strColumn).isNull(),
                    strColumn)
            ).alias(f"Null_Count_For_Column_{strColumn}")
            for strColumn in tblInput.columns
        ]).withColumn("row_id", F.monotonically_increasing_id())
        # Step 2: Get null rate per column
        tblResult2 = tblResult.select(*[
            (F.round(
                F.col(f"Null_Count_For_Column_{strColumn}") / F.lit(intRowCounts),
                4)
            ).alias(f"Null_Rate_For_Column_{strColumn}")
            for strColumn in tblInput.columns
        ]).withColumn("row_id", F.monotonically_increasing_id())
        # Step 3: Combine
        tblResult = tblResult.join(tblResult2,on='row_id',how='inner').drop('row_id')
        return tblResult

    # tested and vetted: does absolutely nothing but checks sets
    def get_set_result(self, tblInput1,strTableName1,strColumnName1,tblInput2,strTableName2,strColumnName2):
        """
        #Inputs:
        1. [tblInput1]: Target table 1.
        2. [strTableName1]: Table name 1.
        3. [strColumnName1]: Column name 1.
        4. [tblInput2]: Target table 2.
        5. [strTableName2]: Table name 2.
        6. [strColumnName2]: Column name 2.

        #Process/Output
        1. Returns a dictionary (`dicResult`) containing the results of the set operations. The function compares the values of specified columns from the two input tables. It computes the following set operations:
        1.1. Intersection: Common elements between the two columns.
        1.2. Difference_1: Elements present in the first column but not in the second.
        1.3. Difference_2: Elements present in the second column but not in the first.
        """
        dicResult = {
            'Intersection':None,
            'Difference_1':None,
            'Difference_2':None,
            'Intersection_Percent':None,
            'Difference_1_Percent':None,
            'Difference_2_Percent':None,
        }
        setInput1 = set([rowRow[strColumnName1] for rowRow in tblInput1.groupBy(strColumnName1).agg(F.count(strColumnName1)).collect()])
        setInput2 = set([rowRow[strColumnName2] for rowRow in tblInput2.groupBy(strColumnName2).agg(F.count(strColumnName2)).collect()])
        dicResult['Intersection'] = setInput1 & setInput2
        dicResult['Difference_1'] = setInput1 - setInput2
        dicResult['Difference_2'] = setInput2 - setInput1
        dicResult['Intersection_Percent'] = len(dicResult['Intersection']) / (len(setInput1) + len(setInput2))
        dicResult['Difference_1_Percent'] = len(dicResult['Difference_1']) / len(setInput1)
        dicResult['Difference_2_Percent'] = len(dicResult['Difference_2']) / len(setInput2)
        return dicResult
    
    # tested and vetted: does absolutely nothing but checks duplicates
    def get_duplicates(self, tblInput, lisstrColumnNamesTarget = None, intMinDuplicateCountFilter = 2):
        """
        #Inputs:
        1. [tblInput]: Target table.
        2. [lisstrColumnNamesTarget]: Optional. List of column names to consider for duplicate identifying.

        #Process/Output
        1. Returns a table containing duplicate rows on the specified list of column names.
        """
        if not lisstrColumnNamesTarget:
            lisstrColumnNamesTarget = tblInput.columns
        tblResultMap = tblInput.groupBy(lisstrColumnNamesTarget)\
                                .agg(F.count('*').alias('count'))\
                                .filter(F.col('count')>=intMinDuplicateCountFilter)
        tblResult = tblInput.join(tblResultMap, on=lisstrColumnNamesTarget, how='inner')
        return tblResult
  
    # outdated, not necessary for newest version, but required for previous eda
    def fix_schema(self,tblInput):

        lisstrColumnNameArray = [strColumnName for strColumnName in tblInput.columns if "collected" in strColumnName.lower()]

        for strColumnName in lisstrColumnNameArray:
            print(f'fixing column: {strColumnName}')
            tblInput = tblInput.withColumn(
                strColumnName,
                F.from_json(F.col(strColumnName), ArrayType(StringType()))
            )
        return tblInput
    
    # outdated, not necessary for newest version, but required for previous eda
    def get_master_table(self,dictblInputs,dicParams):
        """
        #Inputs:
        1. [dictblInputs]: a dictionary of table that will be merged to create the master table. Format of dictionary as follows:
        - dictblInputs = {
            'tblService': tblService,
            'tblVehicle': tblVehicle,
            'tblCustomer': tblCustomer,
            ...
        }
        2. [dicParams]: Optional. Dictionary of parameters.

        #Process/Outputs::
        1. Returns a PySpark table containing the merged table based on predefined joining rules.
        2. Also prints match/fill rate details based on row count
        """
        strColumnName_tblService_CUSTNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_CUSTNO')
        strColumnName_tblCustomer_CUSTNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblCustomer_CUSTNO')
        strColumnName_tblService_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_VEHIID')
        strColumnName_tblVehicle_VEHIID = self.get_parameter_value(dicParams = dicParams, strKey = 'tblVehicle_VEHIID')
        strColumnName_tblService_WRKORDNO = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_WRKORDNO')
        strColumnName_tblService_BILLD = self.get_parameter_value(dicParams = dicParams, strKey = 'tblService_BILLD')
        lisstrColumnNamesPrimaryKeys = self. get_parameter_value(dicParams = dicParams, strKey = 'Primary_Keys')

        lisstrCUSTNOCustomer = [rowRow[strColumnName_tblCustomer_CUSTNO] for rowRow in dictblInputs['tblCustomer'].select(strColumnName_tblCustomer_CUSTNO).distinct().collect()]
        lisstrVEHIIDVehicle = [rowRow[strColumnName_tblVehicle_VEHIID] for rowRow in dictblInputs['tblVehicle'].select(strColumnName_tblVehicle_VEHIID).distinct().collect()]
        lisstrCUSTNOService = [rowRow[strColumnName_tblService_CUSTNO] for rowRow in dictblInputs['tblService'].select(strColumnName_tblService_CUSTNO).distinct().collect()]
        lisstrVEHIIDService = [rowRow[strColumnName_tblService_VEHIID] for rowRow in dictblInputs['tblService'].select(strColumnName_tblService_VEHIID).distinct().collect()]

        print(f"Vehicle to Service Fill Rate:  {len(set(lisstrVEHIIDVehicle))}:{len(set(lisstrVEHIIDService))} || {len(set(lisstrVEHIIDVehicle) & set(lisstrVEHIIDService)) / len(lisstrVEHIIDService)}")
        print(f"Customer to Service Fill Rate: {len(set(lisstrCUSTNOCustomer))}:{len(set(lisstrCUSTNOService))} ||  {len(set(lisstrCUSTNOCustomer) & set(lisstrCUSTNOService)) / len(lisstrCUSTNOService)}")

        tblAlpha = dictblInputs['tblService'].join(dictblInputs['tblVehicle'],
                                                    on=(F.col(strColumnName_tblService_VEHIID) == F.col(strColumnName_tblVehicle_VEHIID)),
                                                    how='left')
        tblBravo = tblAlpha.join(dictblInputs['tblCustomer'],
                                on=(F.col(strColumnName_tblService_CUSTNO) == F.col(strColumnName_tblCustomer_CUSTNO)),
                                how='left')    
        tblResult = tblBravo.select(*lisstrColumnNamesPrimaryKeys,*[strColumnName for strColumnName in tblBravo.columns if strColumnName not in lisstrColumnNamesPrimaryKeys])\
                            .orderBy(F.desc(strColumnName_tblService_WRKORDNO),F.desc(strColumnName_tblService_BILLD))
        return tblResult
    
    def check_count(self,tblInput):
        if tblInput.count()<5000:
            raise('Error! Count of resulting table is too few')